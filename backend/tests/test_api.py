from io import BytesIO


def test_health(client):
    res = client.get("/api/health")
    assert res.status_code == 200
    assert res.json()["status"] == "ok"


def test_login_rejects_bad_password(client):
    res = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
    assert res.status_code == 401


def test_rbi_flow_and_export(client, auth):
    project = client.post(
        "/api/projects",
        headers=auth,
        json={
            "name": "Site A RBI",
            "customer": "Acme",
            "site_name": "DC1",
            "revision": "A",
            "status": "phase2",
            "photography_rules": "No photos in cage 7",
            "restricted_equipment_notes": "EMSS rack R12",
        },
    )
    assert project.status_code == 201, project.text
    pid = project.json()["id"]

    area = client.post(
        f"/api/projects/{pid}/areas",
        headers=auth,
        json={"name": "Hall A", "in_scope": True, "restricted": False},
    )
    assert area.status_code == 201
    aid = area.json()["id"]

    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "area_id": aid, "row_label": "A", "position": "01", "ru_height": 42},
    )
    assert rack.status_code == 201
    rid = rack.json()["id"]

    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "sw-a01",
            "rack_id": rid,
            "vendor": "Cisco",
            "model": "C9300-48P",
            "serial": "FCW1234",
            "device_type": "switch",
            "ru_start": 41,
            "ru_end": 42,
            "fan_orientation": "incorrect-hot-aisle",
            "eol_date": "2020-01-01",
        },
    )
    assert device.status_code == 201, device.text
    did = device.json()["id"]
    assert device.json()["eol_status"] == "eol"

    pdu = client.post(
        f"/api/projects/{pid}/racks/{rid}/pdus",
        headers=auth,
        json={"name": "PDU-A", "bank": "A", "outlet_count": 8, "amperage": 30, "voltage": 208},
    )
    assert pdu.status_code == 201, pdu.text
    pdu_id = pdu.json()["id"]
    port_id = pdu.json()["ports"][0]["id"]
    mapped = client.patch(
        f"/api/projects/{pid}/pdus/{pdu_id}/ports/{port_id}",
        headers=auth,
        json={"port_label": "1", "device_id": did, "notes": "C14"},
    )
    assert mapped.status_code == 200
    assert mapped.json()["ports"][0]["device_id"] == did

    cab = client.post(
        f"/api/projects/{pid}/cables",
        headers=auth,
        json={
            "rack_id": rid,
            "from_label": "sw-a01",
            "from_port": "Gi1/0/1",
            "to_label": "patch-A",
            "to_port": "01",
            "media": "Cat6",
            "traced": True,
        },
    )
    assert cab.status_code == 201

    hand = client.post(
        f"/api/projects/{pid}/handoffs",
        headers=auth,
        json={
            "handoff_date": "2026-08-20",
            "from_name": "Onsite",
            "to_name": "Remote",
            "summary": "Captured A01",
            "devices_captured": 1,
        },
    )
    assert hand.status_code == 201

    lists = client.get(f"/api/projects/{pid}/checklists", headers=auth)
    assert lists.status_code == 200
    assert len(lists.json()) >= 4

    elevation = client.get(f"/api/projects/{pid}/racks/{rid}/elevation", headers=auth)
    assert elevation.status_code == 200
    assert any(s["device_id"] == did for s in elevation.json()["slots"])

    xlsx = client.get(f"/api/projects/{pid}/export.xlsx", headers=auth)
    assert xlsx.status_code == 200
    assert xlsx.content[:2] == b"PK"

    svg = client.get(f"/api/projects/{pid}/racks/{rid}/elevation.svg", headers=auth)
    assert svg.status_code == 200
    assert b"<svg" in svg.content

    dash = client.get("/api/dashboard", headers=auth)
    assert dash.status_code == 200
    body = dash.json()
    assert body["devices"] >= 1
    assert body["eol_devices"] >= 1
    assert body["fan_issues"] >= 1


def test_ops_and_upload(client, auth):
    inc = client.post(
        "/api/incidents",
        headers=auth,
        json={"title": "PDU A breaker trip", "severity": "high", "category": "power", "status": "open"},
    )
    assert inc.status_code == 201
    insp = client.post(
        "/api/inspections",
        headers=auth,
        json={"title": "Daily walkthrough", "itype": "routine", "status": "open", "location": "Hall A"},
    )
    assert insp.status_code == 201
    wo = client.post(
        "/api/work-orders",
        headers=auth,
        json={"title": "Install fiber tray", "wtype": "cabling", "status": "planned"},
    )
    assert wo.status_code == 201
    bp = client.post(
        "/api/backup-processes",
        headers=auth,
        json={"name": "VMware Veeam", "method": "nfs", "schedule": "nightly", "rpo_hours": 24},
    )
    assert bp.status_code == 201

    files = {"file": ("photo.jpg", BytesIO(b"\xff\xd8fakejpeg"), "image/jpeg")}
    up = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "incident", "entity_id": str(inc.json()["id"])},
        files=files,
    )
    assert up.status_code == 201, up.text
    aid = up.json()["id"]
    dl = client.get(f"/api/attachments/{aid}/download", headers=auth)
    assert dl.status_code == 200
    assert dl.content.startswith(b"\xff\xd8")

    backup = client.post("/api/app-backups", headers=auth)
    assert backup.status_code == 200
    assert backup.json()["status"] in ("ok", "error")


def test_catalog_edit_import_search_photos(client, auth):
    catalog = client.get("/api/catalog", headers=auth)
    assert catalog.status_code == 200, catalog.text
    body = catalog.json()
    assert "router" in body["device_types"]
    assert any(t["id"] == "led" for t in body["indicator_types"])
    assert any(c["id"] == "green" for c in body["indicator_colors"])
    names = [v["name"] for v in body["vendors"]]
    for vendor in ("Cisco", "Juniper", "Arista", "MikroTik", "TRENDnet", "Dell", "HPE Aruba", "Fortinet", "Ubiquiti", "Netgear"):
        assert vendor in names
    assert 52 in body["rack_height_presets"]

    project = client.post("/api/projects", headers=auth, json={"name": "Import Site", "customer": "Acme"})
    assert project.status_code == 201, project.text
    pid = project.json()["id"]

    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_label": "A", "ru_height": 42},
    )
    assert rack.status_code == 201
    rid = rack.json()["id"]

    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "sw-typo",
            "rack_id": rid,
            "vendor": "Csico",
            "model": "C9300",
            "serial": "FCW-EDIT",
            "device_type": "switch",
            "ru_start": 47,
            "ru_end": 48,
        },
    )
    assert device.status_code == 201, device.text
    did = device.json()["id"]
    grown = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert any(r["id"] == rid and r["ru_height"] >= 48 for r in grown)

    tall = client.patch(
        f"/api/projects/{pid}/racks/{rid}",
        headers=auth,
        json={"name": "A01", "row_label": "A", "position": "01", "ru_height": 52, "width_inches": 19, "notes": ""},
    )
    assert tall.status_code == 200
    assert tall.json()["ru_height"] == 52

    patched = client.patch(
        f"/api/projects/{pid}/devices/{did}",
        headers=auth,
        json={"vendor": "Cisco", "model": "Catalyst 9300", "device_type": "router", "name": "sw-typo-fixed"},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["vendor"] == "Cisco"
    assert patched.json()["serial"] == "FCW-EDIT"
    assert patched.json()["device_type"] == "router"
    assert patched.json()["name"] == "sw-typo-fixed"

    csv_body = (
        "name,hostname,vendor,model,serial,rack,ru start,height,type,function,management ip\n"
        "core-rtr,core-rtr.site,Cisco,ASR 1001-X,SN-RTR,B12,47,2,router,WAN edge,10.0.0.1\n"
        "logical-fw,fw1,Fortinet,FortiGate 200F,SN-UNLOCATED,,,,firewall,edge,\n"
        "sw-typo-fixed,core-sw,Cisco,Catalyst 9300,FCW-EDIT,A01,50,2,switch,access,\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("gear.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert summary["created"] >= 2
    assert summary["updated"] >= 1
    assert summary["racks_created"] >= 1

    racks = {r["name"]: r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json()}
    assert "B12" in racks
    assert racks["B12"]["ru_height"] >= 48

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Serial", "Vendor", "Model", "Rack", "RU start", "Type"])
    ws.append(["leaf-sw", "SN-XLSX", "Arista", "DCS-7050SX3-48YC8", "B12", 1, "switch"])
    buf = BytesIO()
    wb.save(buf)
    xlsx = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={
            "file": (
                "gear.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert xlsx.status_code == 200, xlsx.text
    assert xlsx.json()["created"] >= 1

    bad = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("legacy.xls", b"\xd0\xcf\x11\xe0", "application/vnd.ms-excel")},
    )
    assert bad.status_code == 400

    search = client.get(f"/api/projects/{pid}/search", headers=auth, params={"q": "FortiGate", "unlocated": True})
    assert search.status_code == 200, search.text
    hits = search.json()["devices"]
    assert any(h["serial"] == "SN-UNLOCATED" for h in hits)
    unlocated_id = next(h["id"] for h in hits if h["serial"] == "SN-UNLOCATED")

    located = client.patch(
        f"/api/projects/{pid}/devices/{unlocated_id}",
        headers=auth,
        json={"rack_id": racks["B12"]["id"], "ru_start": 10, "ru_end": 10},
    )
    assert located.status_code == 200
    assert located.json()["rack_id"] == racks["B12"]["id"]
    assert located.json()["vendor"] == "Fortinet"

    files = {"file": ("faceplate.jpg", BytesIO(b"\xff\xd8one"), "image/jpeg")}
    a1 = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "device", "entity_id": str(did)},
        files=files,
    )
    assert a1.status_code == 201, a1.text
    files2 = {"file": ("rear.jpg", BytesIO(b"\xff\xd8two"), "image/jpeg")}
    a2 = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "device", "entity_id": str(did)},
        files=files2,
    )
    assert a2.status_code == 201
    listed = client.get("/api/attachments", headers=auth, params={"entity_type": "device", "entity_id": did})
    assert listed.status_code == 200
    assert len(listed.json()) >= 2

    got = client.get(f"/api/projects/{pid}/devices/{did}", headers=auth)
    assert got.status_code == 200
    assert got.json()["serial"] == "FCW-EDIT"


def test_auth_required(client):
    res = client.get("/api/projects")
    assert res.status_code == 401


def test_catalog_learns_other_values(client, auth):
    learned = client.post(
        "/api/catalog/learn",
        headers=auth,
        json={"vendor": "CustomOEM", "model": "Widget-9000", "device_type": "blade chassis", "function": "WAN edge"},
    )
    assert learned.status_code == 200, learned.text
    body = learned.json()
    names = [v["name"] for v in body["vendors"]]
    assert "CustomOEM" in names
    oem = next(v for v in body["vendors"] if v["name"] == "CustomOEM")
    assert "Widget-9000" in oem["models"]
    assert "blade chassis" in body["device_types"]
    assert "WAN edge" in body["functions"]

    project = client.post("/api/projects", headers=auth, json={"name": "Catalog persist"})
    pid = project.json()["id"]
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "custom-box", "vendor": "SiteGear", "model": "SG-12", "device_type": "console", "function": "OOB"},
    )
    assert device.status_code == 201, device.text
    catalog = client.get("/api/catalog", headers=auth).json()
    names = [v["name"] for v in catalog["vendors"]]
    assert "SiteGear" in names
    site = next(v for v in catalog["vendors"] if v["name"] == "SiteGear")
    assert "SG-12" in site["models"]
    assert "console" in catalog["device_types"]
    assert "OOB" in catalog["functions"]


def test_admin_can_edit_user_email_and_password(client, auth):
    created = client.post(
        "/api/users",
        headers=auth,
        json={
            "username": "tech1",
            "email": "tech1@example.test",
            "password": "oldpass12",
            "full_name": "Tech One",
            "role": "engineer",
        },
    )
    assert created.status_code == 201, created.text
    uid = created.json()["id"]
    patched = client.patch(
        f"/api/users/{uid}",
        headers=auth,
        json={"email": "tech1.new@example.test", "password": "newpass12", "username": "tech1b", "full_name": "Tech 1"},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["email"] == "tech1.new@example.test"
    assert patched.json()["username"] == "tech1b"
    assert patched.json()["full_name"] == "Tech 1"
    old = client.post("/api/auth/login", json={"username": "tech1b", "password": "oldpass12"})
    assert old.status_code == 401
    new = client.post("/api/auth/login", json={"username": "tech1b", "password": "newpass12"})
    assert new.status_code == 200, new.text


def test_import_preview_mapping_and_rbi_sheet(client, auth):
    from openpyxl import Workbook

    src = client.post("/api/projects", headers=auth, json={"name": "Export site", "customer": "Acme"})
    src_id = src.json()["id"]
    client.post(
        f"/api/projects/{src_id}/devices",
        headers=auth,
        json={"name": "core-sw", "vendor": "Cisco", "model": "Catalyst 9300", "serial": "SN-RBI", "device_type": "switch"},
    )
    exported = client.get(f"/api/projects/{src_id}/export.xlsx", headers=auth)
    assert exported.status_code == 200

    preview = client.post(
        "/api/imports/preview",
        headers=auth,
        files={
            "file": (
                "rbi.xlsx",
                exported.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200, preview.text
    body = preview.json()
    sheet_names = [s["name"] for s in body["sheets"]]
    assert "Cover" in sheet_names
    assert "Devices" in sheet_names
    assert body["suggested_sheet"] == "Devices"
    devices_sheet = next(s for s in body["sheets"] if s["name"] == "Devices")
    assert "name" in devices_sheet["mapped_fields"]
    assert devices_sheet["record_count"] >= 1

    dest = client.post("/api/projects", headers=auth, json={"name": "Import dest"})
    dest_id = dest.json()["id"]
    imported = client.post(
        f"/api/projects/{dest_id}/import",
        headers=auth,
        files={
            "file": (
                "rbi.xlsx",
                exported.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert summary["sheet"] == "Devices"
    assert summary["created"] >= 1
    listed = client.get(f"/api/projects/{dest_id}/devices", headers=auth).json()
    assert any(d["serial"] == "SN-RBI" for d in listed)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gear"
    ws["A1"] = "Name"
    ws["B1"] = "sw-a"
    ws["C1"] = "sw-b"
    ws["A2"] = "Manufacturer"
    ws["B2"] = "Juniper"
    ws["C2"] = "Dell"
    ws["A3"] = "Model"
    ws["B3"] = "EX4400"
    ws["C3"] = "PowerEdge R750"
    ws["A4"] = "Serial"
    ws["B4"] = "SN-COL-A"
    ws["C4"] = "SN-COL-B"
    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    col_preview = client.post(
        "/api/imports/preview",
        headers=auth,
        files={
            "file": (
                "cols.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert col_preview.status_code == 200, col_preview.text
    col_body = col_preview.json()
    assert col_body["sheets"][0]["orientation"] == "columns"
    assert col_body["sheets"][0]["record_count"] == 2

    dest2 = client.post("/api/projects", headers=auth, json={"name": "Column dest"})
    dest2_id = dest2.json()["id"]
    mapped = client.post(
        f"/api/projects/{dest2_id}/import",
        headers=auth,
        data={
            "sheet": "Gear",
            "orientation": "columns",
            "mapping": '{"name": 0, "vendor": 1, "model": 2, "serial": 3}',
        },
        files={
            "file": (
                "cols.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert mapped.status_code == 200, mapped.text
    assert mapped.json()["created"] >= 2
    rows = client.get(f"/api/projects/{dest2_id}/devices", headers=auth).json()
    serials = {d["serial"] for d in rows}
    assert "SN-COL-A" in serials
    assert "SN-COL-B" in serials
    assert any(d["vendor"] == "Juniper" and d["model"] == "EX4400" for d in rows)


def test_rows_relocate_indicators_and_device_copy_move(client, auth):
    src = client.post("/api/projects", headers=auth, json={"name": "Layout src"}).json()
    dest = client.post("/api/projects", headers=auth, json={"name": "Layout dest"}).json()
    src_id, dest_id = src["id"], dest["id"]

    area = client.post(f"/api/projects/{src_id}/areas", headers=auth, json={"name": "Hall A"})
    assert area.status_code == 201
    aid = area.json()["id"]

    row = client.post(f"/api/projects/{src_id}/rows", headers=auth, json={"name": "Row 1", "area_id": aid})
    assert row.status_code == 201
    row_id = row.json()["id"]

    rack = client.post(
        f"/api/projects/{src_id}/racks",
        headers=auth,
        json={"name": "R1", "area_id": aid, "row_id": row_id, "ru_height": 42},
    )
    assert rack.status_code == 201
    rack_id = rack.json()["id"]
    assert rack.json()["row_id"] == row_id
    assert rack.json()["row_label"] == "Row 1"

    auto = client.post(
        f"/api/projects/{src_id}/racks",
        headers=auth,
        json={"name": "R-auto", "area_id": aid, "row_label": "Row Z", "ru_height": 42},
    )
    assert auto.status_code == 201, auto.text
    assert auto.json()["row_id"]
    rows = client.get(f"/api/projects/{src_id}/rows", headers=auth).json()
    assert any(r["name"] == "Row Z" for r in rows)

    hall_b = client.post(f"/api/projects/{src_id}/areas", headers=auth, json={"name": "Hall B"})
    assert hall_b.status_code == 201
    hall_b_id = hall_b.json()["id"]
    reassigned = client.patch(
        f"/api/projects/{src_id}/rows/{row_id}",
        headers=auth,
        json={"name": "Row 1", "area_id": hall_b_id},
    )
    assert reassigned.status_code == 200
    assert reassigned.json()["area_id"] == hall_b_id
    r1 = next(r for r in client.get(f"/api/projects/{src_id}/racks", headers=auth).json() if r["id"] == rack_id)
    assert r1["area_id"] == hall_b_id

    device = client.post(
        f"/api/projects/{src_id}/devices",
        headers=auth,
        json={
            "name": "sw-1",
            "rack_id": rack_id,
            "vendor": "Cisco",
            "serial": "SN-LED",
            "indicator_type": "led",
            "indicator_color": "green",
            "ru_start": 1,
            "ru_end": 1,
        },
    )
    assert device.status_code == 201, device.text
    did = device.json()["id"]
    assert device.json()["indicator_type"] == "led"
    assert device.json()["indicator_color"] == "green"

    patched = client.patch(
        f"/api/projects/{src_id}/devices/{did}",
        headers=auth,
        json={"indicator_type": "both", "indicator_color": "amber"},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["indicator_type"] == "both"
    assert patched.json()["indicator_color"] == "amber"

    copied_area = client.post(
        f"/api/projects/{src_id}/areas/{hall_b_id}/copy",
        headers=auth,
        json={"target_project_id": dest_id, "include_children": True, "include_devices": True},
    )
    assert copied_area.status_code == 200, copied_area.text
    dest_devices = client.get(f"/api/projects/{dest_id}/devices", headers=auth).json()
    assert any(d["serial"] == "SN-LED" and d["id"] != did for d in dest_devices)
    assert any(d["id"] == did for d in client.get(f"/api/projects/{src_id}/devices", headers=auth).json())

    search = client.get(f"/api/projects/{src_id}/search", headers=auth, params={"q": "Row 1"})
    assert search.status_code == 200, search.text
    assert any(h["id"] == did for h in search.json()["devices"])

    dest_area = client.post(f"/api/projects/{dest_id}/areas", headers=auth, json={"name": "Dest Hall"}).json()
    dest_row = client.post(
        f"/api/projects/{dest_id}/rows",
        headers=auth,
        json={"name": "D1", "area_id": dest_area["id"]},
    ).json()
    dest_rack = client.post(
        f"/api/projects/{dest_id}/racks",
        headers=auth,
        json={"name": "DR1", "area_id": dest_area["id"], "row_id": dest_row["id"], "ru_height": 42},
    ).json()

    copied = client.post(
        f"/api/projects/{src_id}/devices/{did}/copy",
        headers=auth,
        json={"target_project_id": dest_id},
    )
    assert copied.status_code == 200, copied.text
    copy_id = copied.json()["id"]
    assert copy_id != did
    assert copied.json()["project_id"] == dest_id
    assert copied.json()["rack_id"] is None
    assert copied.json()["indicator_type"] == "both"
    assert client.get(f"/api/projects/{src_id}/devices/{did}", headers=auth).status_code == 200

    placed = client.post(
        f"/api/projects/{src_id}/devices/{did}/copy",
        headers=auth,
        json={"target_project_id": dest_id, "target_rack_id": dest_rack["id"]},
    )
    assert placed.status_code == 200, placed.text
    assert placed.json()["rack_id"] == dest_rack["id"]
    assert placed.json()["id"] != did
    assert client.get(f"/api/projects/{src_id}/devices/{did}", headers=auth).status_code == 200

    bad_rack = client.post(
        f"/api/projects/{src_id}/devices/{did}/copy",
        headers=auth,
        json={"target_project_id": dest_id, "target_rack_id": rack_id},
    )
    assert bad_rack.status_code == 404

    moved = client.post(
        f"/api/projects/{src_id}/devices/{did}/move",
        headers=auth,
        json={"target_project_id": dest_id, "target_rack_id": dest_rack["id"]},
    )
    assert moved.status_code == 200, moved.text
    assert moved.json()["project_id"] == dest_id
    assert moved.json()["rack_id"] == dest_rack["id"]
    assert client.get(f"/api/projects/{src_id}/devices/{did}", headers=auth).status_code == 404
    assert client.get(f"/api/projects/{dest_id}/devices/{did}", headers=auth).status_code == 200

    moved_rack = client.post(
        f"/api/projects/{src_id}/racks/{rack_id}/move",
        headers=auth,
        json={
            "target_project_id": dest_id,
            "target_area_id": dest_area["id"],
            "target_row_id": dest_row["id"],
            "include_devices": True,
        },
    )
    assert moved_rack.status_code == 200, moved_rack.text
    assert moved_rack.json()["project_id"] == dest_id
    assert moved_rack.json()["row_id"] == dest_row["id"]
    src_racks = client.get(f"/api/projects/{src_id}/racks", headers=auth).json()
    dest_racks = client.get(f"/api/projects/{dest_id}/racks", headers=auth).json()
    assert all(r["id"] != rack_id for r in src_racks)
    assert any(r["id"] == rack_id for r in dest_racks)



def test_area_row_rack_hierarchy_copy_move_search(client, auth):
    src = client.post("/api/projects", headers=auth, json={"name": "Layout src", "customer": "Acme"})
    assert src.status_code == 201, src.text
    pid = src.json()["id"]
    dest = client.post("/api/projects", headers=auth, json={"name": "Layout dest"})
    dest_id = dest.json()["id"]

    hall = client.post(
        f"/api/projects/{pid}/areas",
        headers=auth,
        json={"name": "Hall A", "in_scope": True},
    )
    assert hall.status_code == 201, hall.text
    hall_id = hall.json()["id"]
    hall_b = client.post(
        f"/api/projects/{pid}/areas",
        headers=auth,
        json={"name": "Hall B"},
    )
    hall_b_id = hall_b.json()["id"]

    row = client.post(
        f"/api/projects/{pid}/rows",
        headers=auth,
        json={"name": "Row 1", "area_id": hall_id},
    )
    assert row.status_code == 201, row.text
    row_id = row.json()["id"]
    assert row.json()["area_id"] == hall_id

    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_id": row_id, "area_id": hall_id, "ru_height": 42},
    )
    assert rack.status_code == 201, rack.text
    rid = rack.json()["id"]
    assert rack.json()["row_id"] == row_id
    assert rack.json()["row_label"] == "Row 1"
    assert rack.json()["area_id"] == hall_id

    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "sw-a01", "rack_id": rid, "vendor": "Cisco", "serial": "SN-ROW", "device_type": "switch"},
    )
    assert device.status_code == 201, device.text

    reassigned = client.patch(
        f"/api/projects/{pid}/rows/{row_id}",
        headers=auth,
        json={"name": "Row 1", "area_id": hall_b_id},
    )
    assert reassigned.status_code == 200, reassigned.text
    assert reassigned.json()["area_id"] == hall_b_id
    moved_rack = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert any(r["id"] == rid and r["area_id"] == hall_b_id for r in moved_rack)

    copied = client.post(
        f"/api/projects/{pid}/areas/{hall_b_id}/copy",
        headers=auth,
        json={"target_project_id": dest_id, "include_children": True, "include_devices": False},
    )
    assert copied.status_code == 200, copied.text
    dest_areas = client.get(f"/api/projects/{dest_id}/areas", headers=auth).json()
    dest_rows = client.get(f"/api/projects/{dest_id}/rows", headers=auth).json()
    dest_racks = client.get(f"/api/projects/{dest_id}/racks", headers=auth).json()
    dest_devices = client.get(f"/api/projects/{dest_id}/devices", headers=auth).json()
    assert any(a["name"] == "Hall B" for a in dest_areas)
    assert any(r["name"] == "Row 1" for r in dest_rows)
    assert any(r["name"] == "A01" for r in dest_racks)
    assert dest_devices == []

    still_src = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert any(r["id"] == rid for r in still_src)

    dest_rows = client.get(f"/api/projects/{dest_id}/rows", headers=auth).json()
    dest_row_id = next(r["id"] for r in dest_rows if r["name"] == "Row 1")
    moved = client.post(
        f"/api/projects/{pid}/racks/{rid}/move",
        headers=auth,
        json={"target_project_id": dest_id, "target_row_id": dest_row_id, "include_devices": True},
    )
    assert moved.status_code == 200, moved.text
    assert moved.json()["project_id"] == dest_id
    src_racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert all(r["id"] != rid for r in src_racks)
    dest_racks = client.get(f"/api/projects/{dest_id}/racks", headers=auth).json()
    assert any(r["id"] == rid for r in dest_racks)
    dest_devices = client.get(f"/api/projects/{dest_id}/devices", headers=auth).json()
    assert any(d["serial"] == "SN-ROW" for d in dest_devices)

    search = client.get(f"/api/projects/{dest_id}/search", headers=auth, params={"q": "Row 1"})
    assert search.status_code == 200, search.text
    hits = search.json()["devices"]
    assert any(h["serial"] == "SN-ROW" for h in hits)
    hit = next(h for h in hits if h["serial"] == "SN-ROW")
    assert hit["rack_row"] == "Row 1"
    assert hit["area_name"]

    auto = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "B02", "row_label": "AutoRow", "ru_height": 42},
    )
    assert auto.status_code == 201, auto.text
    assert auto.json()["row_id"]
    listed = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    assert any(r["name"] == "AutoRow" and r["id"] == auto.json()["row_id"] for r in listed)


def test_hierarchy_delete_unassigns_children(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Delete site"})
    assert project.status_code == 201, project.text
    pid = project.json()["id"]
    area = client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Hall Del"}).json()
    row = client.post(
        f"/api/projects/{pid}/rows",
        headers=auth,
        json={"name": "Row Del", "area_id": area["id"]},
    ).json()
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "D01", "row_id": row["id"], "area_id": area["id"], "ru_height": 42},
    ).json()
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "sw-del", "rack_id": rack["id"], "serial": "SN-DEL", "ru_start": 10, "ru_end": 10},
    ).json()

    gone_device = client.delete(f"/api/projects/{pid}/devices/{device['id']}", headers=auth)
    assert gone_device.status_code == 200
    assert client.get(f"/api/projects/{pid}/devices/{device['id']}", headers=auth).status_code == 404

    device2 = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "sw-del-2", "rack_id": rack["id"], "serial": "SN-DEL-2", "ru_start": 11, "ru_end": 11},
    )
    assert device2.status_code == 201, device2.text
    did2 = device2.json()["id"]
    gone_rack = client.delete(f"/api/projects/{pid}/racks/{rack['id']}", headers=auth)
    assert gone_rack.status_code == 200
    leftover = client.get(f"/api/projects/{pid}/devices/{did2}", headers=auth)
    assert leftover.status_code == 200
    assert leftover.json()["rack_id"] is None
    assert leftover.json()["serial"] == "SN-DEL-2"

    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "D02", "row_id": row["id"], "area_id": area["id"], "ru_height": 42},
    ).json()
    gone_row = client.delete(f"/api/projects/{pid}/rows/{row['id']}", headers=auth)
    assert gone_row.status_code == 200
    kept_rack = next(r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json() if r["id"] == rack["id"])
    assert kept_rack["row_id"] is None
    assert client.get(f"/api/projects/{pid}/rows", headers=auth).json() == []

    gone_area = client.delete(f"/api/projects/{pid}/areas/{area['id']}", headers=auth)
    assert gone_area.status_code == 200, gone_area.text
    assert client.get(f"/api/projects/{pid}/areas", headers=auth).json() == []
    kept_rack = next(r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json() if r["id"] == rack["id"])
    assert kept_rack["area_id"] is None


def _login(client, username, password):
    res = client.post("/api/auth/login", json={"username": username, "password": password})
    assert res.status_code == 200, res.text
    return {"Authorization": f"Bearer {res.json()['access_token']}"}


def _create_user(client, auth, username, role, password="engineer12"):
    res = client.post(
        "/api/users",
        headers=auth,
        json={
            "username": username,
            "email": f"{username}@example.test",
            "password": password,
            "full_name": username,
            "role": role,
        },
    )
    assert res.status_code == 201, res.text
    return res.json()


def _ods_bytes(headers, rows, sheet="Layout"):
    import zipfile
    from xml.etree.ElementTree import Element, SubElement, tostring

    ns = {
        "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
        "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
        "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    }

    def q(prefix, tag):
        return f"{{{ns[prefix]}}}{tag}"

    root = Element(q("office", "document-content"))
    for prefix, uri in ns.items():
        root.set(f"xmlns:{prefix}", uri)
    body = SubElement(root, q("office", "body"))
    spreadsheet = SubElement(body, q("office", "spreadsheet"))
    table = SubElement(spreadsheet, q("table", "table"))
    table.set(q("table", "name"), sheet)
    for values in [headers, *rows]:
        row_el = SubElement(table, q("table", "table-row"))
        for value in values:
            cell = SubElement(row_el, q("table", "table-cell"))
            p = SubElement(cell, q("text", "p"))
            p.text = str(value)
    content = tostring(root, encoding="utf-8", xml_declaration=True)
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("mimetype", "application/vnd.oasis.opendocument.spreadsheet", compress_type=zipfile.ZIP_STORED)
        zf.writestr("content.xml", content)
    return buf.getvalue()


def test_admin_rename_and_delete_project(client, auth):
    created = client.post("/api/projects", headers=auth, json={"name": "Doomed site", "customer": "Acme"})
    assert created.status_code == 201, created.text
    pid = created.json()["id"]
    client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Hall X"})
    renamed = client.patch(
        f"/api/projects/{pid}",
        headers=auth,
        json={"name": "Renamed site", "customer": "Acme"},
    )
    assert renamed.status_code == 200, renamed.text
    assert renamed.json()["name"] == "Renamed site"
    gone = client.delete(f"/api/projects/{pid}", headers=auth)
    assert gone.status_code == 200, gone.text
    assert client.get(f"/api/projects/{pid}", headers=auth).status_code == 404
    names = [p["name"] for p in client.get("/api/projects", headers=auth).json()]
    assert "Renamed site" not in names
    assert "Doomed site" not in names


def test_engineer_cannot_rename_or_delete_project_but_can_import(client, auth):
    _create_user(client, auth, "eng-layout", "engineer")
    _create_user(client, auth, "remote-layout", "remote")
    eng = _login(client, "eng-layout", "engineer12")
    remote = _login(client, "remote-layout", "engineer12")

    project = client.post("/api/projects", headers=eng, json={"name": "Eng site"})
    assert project.status_code == 201, project.text
    pid = project.json()["id"]

    blocked_rename = client.patch(f"/api/projects/{pid}", headers=eng, json={"name": "Hacked name"})
    assert blocked_rename.status_code == 403
    still = client.get(f"/api/projects/{pid}", headers=eng).json()
    assert still["name"] == "Eng site"

    blocked_delete = client.delete(f"/api/projects/{pid}", headers=eng)
    assert blocked_delete.status_code == 403
    assert client.get(f"/api/projects/{pid}", headers=eng).status_code == 200

    csv_body = (
        "area,aisle,rack,name,serial\n"
        "Hall A,Row 1,A01,sw-a,SN-HALL\n"
        "Hall A,Row 2,,,\n"
        "Hall A,Row 3,A03,,\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=eng,
        files={"file": ("hall.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert summary["areas_created"] >= 1
    assert summary["rows_created"] >= 3
    assert summary["racks_created"] >= 2
    assert summary["created"] >= 1
    areas = client.get(f"/api/projects/{pid}/areas", headers=eng).json()
    rows = client.get(f"/api/projects/{pid}/rows", headers=eng).json()
    racks = client.get(f"/api/projects/{pid}/racks", headers=eng).json()
    assert any(a["name"] == "Hall A" for a in areas)
    assert {r["name"] for r in rows} >= {"Row 1", "Row 2", "Row 3"}
    assert {r["name"] for r in racks} >= {"A01", "A03"}
    hall_id = next(a["id"] for a in areas if a["name"] == "Hall A")
    assert all(r["area_id"] == hall_id for r in rows if r["name"].startswith("Row"))

    remote_import = client.post(
        f"/api/projects/{pid}/import",
        headers=remote,
        files={"file": ("hall.csv", csv_body.encode(), "text/csv")},
    )
    assert remote_import.status_code == 403
    remote_preview = client.post(
        "/api/imports/preview",
        headers=remote,
        files={"file": ("hall.csv", csv_body.encode(), "text/csv")},
    )
    assert remote_preview.status_code == 403


def test_import_ods_and_default_area_layout(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "ODS site"})
    pid = project.json()["id"]
    area = client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Cage 7"}).json()
    ods = _ods_bytes(
        ["Aisle", "Rack", "Name", "Serial"],
        [
            ["Aisle 1", "C01", "leaf-1", "SN-ODS-1"],
            ["Aisle 2", "", "", ""],
            ["Aisle 1", "C02", "", ""],
        ],
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        data={"default_area_id": str(area["id"])},
        files={"file": ("floor.ods", ods, "application/vnd.oasis.opendocument.spreadsheet")},
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert summary["rows_created"] >= 2
    assert summary["racks_created"] >= 2
    assert summary["created"] >= 1
    rows = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert any(r["name"] == "Aisle 1" and r["area_id"] == area["id"] for r in rows)
    assert any(r["name"] == "Aisle 2" and r["area_id"] == area["id"] for r in rows)
    assert {r["name"] for r in racks} >= {"C01", "C02"}
    assert all(r["area_id"] == area["id"] for r in racks if r["name"] in {"C01", "C02"})


def _xlsx_bytes(sheets):
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for title, headers, rows in sheets:
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        ws.append(headers)
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_does_not_steal_populated_rows_across_areas(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Hierarchy halls"}).json()
    pid = project["id"]
    hall_a = (
        "area,aisle,rack,name,serial,vendor,model\n"
        "Hall A,Row 1,A01,core-a,SN-HALL-A,Cisco,C9300\n"
        "Hall A,Row 1,A02,leaf-a,SN-LEAF-A,Arista,7050\n"
    )
    first = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("hall-a.csv", hall_a.encode(), "text/csv")},
    )
    assert first.status_code == 200, first.text
    areas = {a["name"]: a for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()}
    assert "Hall A" in areas
    rows = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    row_a = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall A"]["id"])
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    hall_a_racks = {r["name"]: r for r in racks if r["row_id"] == row_a["id"]}
    assert set(hall_a_racks) >= {"A01", "A02"}
    a01_id = hall_a_racks["A01"]["id"]
    core = next(d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json() if d["serial"] == "SN-HALL-A")
    assert core["rack_id"] == a01_id
    assert core["vendor"] == "Cisco"

    hall_b = (
        "area,aisle,rack,name,serial,vendor,model\n"
        "Hall B,Row 1,A01,core-b,SN-HALL-B,Juniper,EX4400\n"
        "Hall B,Row 1,A02,leaf-b,SN-LEAF-B,Dell,R750\n"
    )
    second = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("hall-b.csv", hall_b.encode(), "text/csv")},
    )
    assert second.status_code == 200, second.text
    areas = {a["name"]: a for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()}
    assert "Hall A" in areas and "Hall B" in areas
    rows = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    row_a_again = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall A"]["id"])
    row_b = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall B"]["id"])
    assert row_a_again["id"] == row_a["id"]
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    hall_a_racks = [r for r in racks if r["row_id"] == row_a["id"]]
    hall_b_racks = [r for r in racks if r["row_id"] == row_b["id"]]
    assert {r["name"] for r in hall_a_racks} >= {"A01", "A02"}
    assert {r["name"] for r in hall_b_racks} >= {"A01", "A02"}
    assert {r["id"] for r in hall_a_racks}.isdisjoint({r["id"] for r in hall_b_racks})
    devices = {d["serial"]: d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json()}
    assert devices["SN-HALL-A"]["rack_id"] == a01_id
    assert devices["SN-HALL-A"]["vendor"] == "Cisco"
    b01 = next(r for r in hall_b_racks if r["name"] == "A01")
    assert devices["SN-HALL-B"]["rack_id"] == b01["id"]


def test_import_all_sheets_named_halls_without_stealing(client, auth):
    xlsx = _xlsx_bytes(
        [
            ("Hall A", ["Aisle", "Rack", "Name", "Serial", "Vendor"], [["Row 1", "A01", "sw-a", "SN-SHEET-A", "Cisco"]]),
            ("Hall B", ["Aisle", "Rack", "Name", "Serial", "Vendor"], [["Row 1", "A01", "sw-b", "SN-SHEET-B", "Juniper"]]),
        ]
    )
    project = client.post("/api/projects", headers=auth, json={"name": "Multi sheet halls"}).json()
    pid = project["id"]

    default_import = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={
            "file": (
                "halls.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert default_import.status_code == 200, default_import.text
    assert default_import.json()["sheet"] == "Hall A"
    areas = [a["name"] for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()]
    assert "Hall A" in areas
    assert "Hall B" not in areas

    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        data={"all_sheets": "true"},
        files={
            "file": (
                "halls.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert "Hall A" in summary["sheet"] and "Hall B" in summary["sheet"]
    areas = {a["name"]: a for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()}
    assert "Hall A" in areas and "Hall B" in areas
    rows = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    row_a = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall A"]["id"])
    row_b = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall B"]["id"])
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    a01_a = next(r for r in racks if r["name"] == "A01" and r["row_id"] == row_a["id"])
    a01_b = next(r for r in racks if r["name"] == "A01" and r["row_id"] == row_b["id"])
    assert a01_a["id"] != a01_b["id"]
    devices = {d["serial"]: d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json()}
    assert devices["SN-SHEET-A"]["rack_id"] == a01_a["id"]
    assert devices["SN-SHEET-B"]["rack_id"] == a01_b["id"]


def test_import_preserves_serial_in_other_location(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Preserve serial"}).json()
    pid = project["id"]
    first = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={
            "file": (
                "hall-a.csv",
                b"area,aisle,rack,name,serial,vendor\nHall A,Row 1,A01,core-a,SN-SHARED,Cisco\n",
                "text/csv",
            )
        },
    )
    assert first.status_code == 200, first.text
    original = next(d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json() if d["serial"] == "SN-SHARED")
    original_rack = original["rack_id"]
    second = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={
            "file": (
                "hall-b.csv",
                b"area,aisle,rack,name,serial,vendor\nHall B,Row 1,A01,core-b,SN-SHARED,Juniper\n",
                "text/csv",
            )
        },
    )
    assert second.status_code == 200, second.text
    summary = second.json()
    assert summary["preserved"] >= 1
    assert summary["created"] == 0
    devices = [d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json() if d["serial"] == "SN-SHARED"]
    assert len(devices) == 1
    kept = devices[0]
    assert kept["id"] == original["id"]
    assert kept["rack_id"] == original_rack
    assert kept["vendor"] == "Cisco"
    assert kept["name"] == "core-a"
    areas = {a["name"]: a for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()}
    assert "Hall B" in areas
    rows = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    row_b = next(r for r in rows if r["name"] == "Row 1" and r["area_id"] == areas["Hall B"]["id"])
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    assert any(r["name"] == "A01" and r["row_id"] == row_b["id"] for r in racks)


def test_import_empty_cells_do_not_blank_device_fields(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Keep fields"}).json()
    pid = project["id"]
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_label": "A", "ru_height": 42},
    ).json()
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "core-keep",
            "rack_id": rack["id"],
            "hostname": "old-host",
            "vendor": "Cisco",
            "model": "C9300",
            "serial": "SN-KEEP",
            "function": "access",
            "notes": "keep me",
            "fan_orientation": "front-to-back",
            "indicator_type": "led",
            "indicator_color": "green",
        },
    ).json()
    csv_body = (
        "name,serial,rack,vendor,model,notes,function,hostname,fan_orientation,type\n"
        "core-keep,SN-KEEP,A01,,,,,core-keep-host,unknown,\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("blank.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["updated"] >= 1
    got = client.get(f"/api/projects/{pid}/devices/{device['id']}", headers=auth).json()
    assert got["vendor"] == "Cisco"
    assert got["model"] == "C9300"
    assert got["notes"] == "keep me"
    assert got["function"] == "access"
    assert got["fan_orientation"] == "front-to-back"
    assert got["indicator_type"] == "led"
    assert got["indicator_color"] == "green"
    assert got["hostname"] == "core-keep-host"
    assert got["rack_id"] == rack["id"]


def test_import_parses_device_name_into_blank_identity_fields(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Name parse"}).json()
    pid = project["id"]
    csv_body = (
        "name,serial,rack\n"
        "Cisco Catalyst 3560G Ethernet Switch,SN-CAT,A01\n"
        "Floor Widget 12,SN-UNK,A01\n"
        "Cisco Catalyst 9300-48P Switch,SN-9300,A01\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("names.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    devices = {d["serial"]: d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json()}
    cat = devices["SN-CAT"]
    assert cat["vendor"] == "Cisco"
    assert cat["model"] == "Catalyst 3560G"
    assert cat["device_type"] == "switch"
    unk = devices["SN-UNK"]
    assert unk["vendor"] == ""
    assert unk["model"] == ""
    assert unk["device_type"] == ""
    known = devices["SN-9300"]
    assert known["vendor"] == "Cisco"
    assert known["model"] == "Catalyst 9300-48P"
    assert known["device_type"] == "switch"


def test_import_does_not_overwrite_provided_or_existing_identity(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Keep identity"}).json()
    pid = project["id"]
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_label": "A", "ru_height": 42},
    ).json()
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "core-keep",
            "rack_id": rack["id"],
            "vendor": "Juniper",
            "model": "EX4300",
            "serial": "SN-ID",
            "device_type": "switch",
        },
    ).json()
    csv_body = (
        "name,serial,rack,vendor,model,type\n"
        "Cisco Catalyst 3560G Ethernet Switch,SN-ID,A01,,,\n"
        "core-new,SN-NEW,A01,Arista,,firewall\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("keep.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    got = client.get(f"/api/projects/{pid}/devices/{device['id']}", headers=auth).json()
    assert got["vendor"] == "Juniper"
    assert got["model"] == "EX4300"
    assert got["device_type"] == "switch"
    created = next(d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json() if d["serial"] == "SN-NEW")
    assert created["vendor"] == "Arista"
    assert created["model"] == ""
    assert created["device_type"] == "firewall"


def test_device_power_draw_unit_roundtrip(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Power"}).json()
    pid = project["id"]
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "pdu-1", "power_draw_watts": 2500, "power_draw_unit": "kW", "device_type": "pdu"},
    ).json()
    assert device["power_draw_watts"] == 2500
    assert device["power_draw_unit"] == "kW"
    patched = client.patch(
        f"/api/projects/{pid}/devices/{device['id']}",
        headers=auth,
        json={"power_draw_watts": 400, "power_draw_unit": "W"},
    ).json()
    assert patched["power_draw_watts"] == 400
    assert patched["power_draw_unit"] == "W"


def test_device_ac_dc_power_and_dual_pdus(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Feeds"}).json()
    pid = project["id"]
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_label": "A", "ru_height": 42},
    ).json()
    pdu_a = client.post(
        f"/api/projects/{pid}/racks/{rack['id']}/pdus",
        headers=auth,
        json={"name": "PDU-A", "bank": "A", "outlet_count": 4},
    ).json()
    pdu_b = client.post(
        f"/api/projects/{pid}/racks/{rack['id']}/pdus",
        headers=auth,
        json={"name": "PDU-B", "bank": "B", "outlet_count": 4},
    ).json()
    listed = client.get(f"/api/projects/{pid}/pdus", headers=auth)
    assert listed.status_code == 200, listed.text
    assert {p["name"] for p in listed.json()} == {"PDU-A", "PDU-B"}
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "core-sw",
            "rack_id": rack["id"],
            "power_draw_watts": 1800,
            "power_draw_unit": "kW",
            "dc_power_draw_amps": 12.5,
            "pdu_a_id": pdu_a["id"],
            "pdu_b_id": pdu_b["id"],
        },
    ).json()
    assert device["power_draw_watts"] == 1800
    assert device["dc_power_draw_amps"] == 12.5
    assert device["pdu_a_id"] == pdu_a["id"]
    assert device["pdu_b_id"] == pdu_b["id"]
    cleared = client.patch(
        f"/api/projects/{pid}/devices/{device['id']}",
        headers=auth,
        json={"dc_power_draw_amps": 7, "pdu_b_id": None},
    ).json()
    assert cleared["dc_power_draw_amps"] == 7
    assert cleared["pdu_a_id"] == pdu_a["id"]
    assert cleared["pdu_b_id"] is None
    bad = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "ghost", "pdu_a_id": 999999},
    )
    assert bad.status_code == 400


def test_device_owner_create_patch_search_and_export(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Shared cage"}).json()
    pid = project["id"]
    created = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "cust-fw", "vendor": "Fortinet", "serial": "SN-OWN", "owner": "Acme Colo"},
    )
    assert created.status_code == 201, created.text
    assert created.json()["owner"] == "Acme Colo"
    patched = client.patch(
        f"/api/projects/{pid}/devices/{created.json()['id']}",
        headers=auth,
        json={"owner": "Beta Tenant"},
    )
    assert patched.status_code == 200
    assert patched.json()["owner"] == "Beta Tenant"
    search = client.get(f"/api/projects/{pid}/search", headers=auth, params={"q": "Beta"})
    assert search.status_code == 200
    assert any(d["serial"] == "SN-OWN" for d in search.json()["devices"])
    exported = client.get(f"/api/projects/{pid}/export.xlsx", headers=auth)
    assert exported.status_code == 200
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(exported.content), data_only=True)
    devices_sheet = wb["Devices"]
    headers = [cell.value for cell in next(devices_sheet.iter_rows(min_row=1, max_row=1))]
    assert "Owner" in headers
    owner_col = headers.index("Owner") + 1
    owners = [devices_sheet.cell(row, owner_col).value for row in range(2, devices_sheet.max_row + 1)]
    assert "Beta Tenant" in owners


def test_import_parses_location_code_and_owner(client, auth):
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"
    ws.append(["Name", "Serial", "Location", "Owner", "Vendor"])
    ws.append(["edge-fw", "SN-LOC-1", "A12 R09-RU19", "Acme Colo", "Fortinet"])
    ws.append(["core-sw", "SN-LOC-2", "A12 R09-RU20-RU21", "Beta Tenant", "Cisco"])
    buf = BytesIO()
    wb.save(buf)
    project = client.post("/api/projects", headers=auth, json={"name": "Location import"}).json()
    pid = project["id"]

    preview = client.post(
        "/api/imports/preview",
        headers=auth,
        files={
            "file": (
                "gear.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200, preview.text
    sheet = preview.json()["sheets"][0]
    assert "location" in sheet["mapped_fields"]
    assert "owner" in sheet["mapped_fields"]
    sample = sheet["sample_records"][0]
    assert sample["row"] == "A12"
    assert sample["rack"] == "09"
    assert sample["ru_start"] == "19"
    assert sample["owner"] == "Acme Colo"

    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={
            "file": (
                "gear.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["created"] >= 2
    rows = {r["name"]: r for r in client.get(f"/api/projects/{pid}/rows", headers=auth).json()}
    assert "A12" in rows
    racks = client.get(f"/api/projects/{pid}/racks", headers=auth).json()
    rack = next(r for r in racks if r["name"] == "09" and r["row_id"] == rows["A12"]["id"])
    devices = {d["serial"]: d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json()}
    first = devices["SN-LOC-1"]
    assert first["rack_id"] == rack["id"]
    assert first["ru_start"] == 19
    assert first["owner"] == "Acme Colo"
    second = devices["SN-LOC-2"]
    assert second["rack_id"] == rack["id"]
    assert second["ru_start"] == 20
    assert second["ru_end"] == 21
    assert second["owner"] == "Beta Tenant"


def test_import_explicit_rack_column_wins_over_parsed_location(client, auth):
    csv_body = (
        "name,serial,location,rack,owner\n"
        "leaf-sw,SN-OVR,A12 R09-RU19,CAB-1,Acme Colo\n"
    )
    project = client.post("/api/projects", headers=auth, json={"name": "Location override"}).json()
    pid = project["id"]
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("override.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    rows = {r["name"]: r for r in client.get(f"/api/projects/{pid}/rows", headers=auth).json()}
    assert "A12" in rows
    racks = {r["name"]: r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json()}
    assert "CAB-1" in racks
    assert "09" not in racks
    device = next(d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json() if d["serial"] == "SN-OVR")
    assert device["rack_id"] == racks["CAB-1"]["id"]
    assert device["ru_start"] == 19
    assert device["owner"] == "Acme Colo"


def test_visio_office_export_preserves_hierarchy_and_pictures(client, auth):
    from io import BytesIO
    from zipfile import ZipFile

    from openpyxl import load_workbook

    project = client.post(
        "/api/projects",
        headers=auth,
        json={"name": "Azure DC", "customer": "Acme", "site_name": "Hall 3"},
    ).json()
    pid = project["id"]
    area = client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Hall A"}).json()
    row = client.post(f"/api/projects/{pid}/rows", headers=auth, json={"name": "A12", "area_id": area["id"]}).json()
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "09", "area_id": area["id"], "row_id": row["id"], "ru_height": 42},
    ).json()
    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "edge-fw",
            "rack_id": rack["id"],
            "vendor": "Fortinet",
            "serial": "SN-VISIO",
            "owner": "Acme Colo",
            "ru_start": 19,
            "ru_end": 19,
            "device_type": "firewall",
        },
    ).json()
    unlocated = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "spare-sw", "serial": "SN-LOOSE", "device_type": "switch"},
    ).json()
    jpeg = BytesIO(b"\xff\xd8visio-photo")
    up = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "device", "entity_id": str(device["id"])},
        files={"file": ("faceplate.jpg", jpeg, "image/jpeg")},
    )
    assert up.status_code == 201, up.text
    restricted = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "device", "entity_id": str(device["id"]), "photography_restricted": "true"},
        files={"file": ("secret.jpg", BytesIO(b"\xff\xd8secret"), "image/jpeg")},
    )
    assert restricted.status_code == 201, restricted.text
    rack_photo = client.post(
        "/api/attachments",
        headers=auth,
        data={"entity_type": "rack", "entity_id": str(rack["id"])},
        files={"file": ("rack-front.jpg", BytesIO(b"\xff\xd8rack"), "image/jpeg")},
    )
    assert rack_photo.status_code == 201, rack_photo.text

    exported = client.get(f"/api/projects/{pid}/export-visio.zip", headers=auth)
    assert exported.status_code == 200, exported.text
    assert exported.content[:2] == b"PK"
    zf = ZipFile(BytesIO(exported.content))
    names = zf.namelist()
    assert "How to open in Visio.txt" in names
    vsdx_name = next(n for n in names if n.endswith(".vsdx"))
    xlsx_name = next(n for n in names if n.endswith(".xlsx") and "Data Visualizer" in n)
    picture_names = [n for n in names if n.startswith("Pictures/") and n.endswith(".jpg")]
    assert any("Hall A" in n and "A12" in n and "09" in n and "edge-fw" in n for n in picture_names)
    assert any("rack-front.jpg" in n for n in picture_names)
    assert not any("secret.jpg" in n for n in names)
    assert any(n.startswith("Elevations/") and n.endswith("09.svg") for n in names)

    vsdx = ZipFile(BytesIO(zf.read(vsdx_name)))
    vsdx_names = vsdx.namelist()
    assert "visio/document.xml" in vsdx_names
    assert "visio/pages/pages.xml" in vsdx_names
    pages_xml = vsdx.read("visio/pages/pages.xml").decode("utf-8")
    assert "Overview" in pages_xml
    assert "Hall A" in pages_xml
    assert "09" in pages_xml
    assert any(n.startswith("visio/media/") for n in vsdx_names)

    wb = load_workbook(BytesIO(zf.read(xlsx_name)))
    assert "VisioHierarchy" in wb.sheetnames
    vis = wb["VisioHierarchy"]
    rows = list(vis.iter_rows(min_row=2, values_only=True))
    by_id = {r[0]: r for r in rows}
    assert "SITE" in by_id
    area_row = next(r for r in rows if r[2] == "Area")
    row_row = next(r for r in rows if r[2] == "Row")
    rack_row = next(r for r in rows if r[2] == "Rack")
    device_row = next(r for r in rows if r[0].startswith("DEV-") and r[1] == "edge-fw")
    assert area_row[3] == "SITE"
    assert row_row[3] == area_row[0]
    assert rack_row[3] == row_row[0]
    assert device_row[3] == rack_row[0]
    assert device_row[9] == "Acme Colo"
    assert "faceplate.jpg" in (device_row[12] or "")
    assert any(r[1] == "spare-sw" for r in rows)
    assert unlocated["id"]


def test_bulk_create_rows_under_area(client, auth):
    project = client.post(
        "/api/projects",
        headers=auth,
        json={"name": "Row Set", "customer": "Acme", "site_name": "DC1", "revision": "A"},
    )
    pid = project.json()["id"]
    area = client.post(
        f"/api/projects/{pid}/areas",
        headers=auth,
        json={"name": "Hall A", "in_scope": True},
    )
    aid = area.json()["id"]
    empty = client.post(f"/api/projects/{pid}/rows/bulk", headers=auth, json={"area_id": aid, "names": ["  ", ""]})
    assert empty.status_code == 400
    created = client.post(
        f"/api/projects/{pid}/rows/bulk",
        headers=auth,
        json={"area_id": aid, "names": ["A01", "A02", "a01", "  A03  "]},
    )
    assert created.status_code == 201, created.text
    body = created.json()
    assert [r["name"] for r in body["created"]] == ["A01", "A02", "A03"]
    assert body["existing"] == []
    again = client.post(
        f"/api/projects/{pid}/rows/bulk",
        headers=auth,
        json={"area_id": aid, "names": ["A01", "A04"]},
    )
    assert again.status_code == 201
    assert [r["name"] for r in again.json()["created"]] == ["A04"]
    assert [r["name"] for r in again.json()["existing"]] == ["A01"]
    listed = client.get(f"/api/projects/{pid}/rows", headers=auth).json()
    assert [r["name"] for r in listed] == ["A01", "A02", "A03", "A04"]
    assert all(r["area_id"] == aid for r in listed)


def test_new_device_type_defaults_blank(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Blank type"}).json()
    pid = project["id"]
    device = client.post(f"/api/projects/{pid}/devices", headers=auth, json={"name": "new-box"}).json()
    assert device["device_type"] == ""
    listed = client.get(f"/api/projects/{pid}/devices", headers=auth).json()
    assert listed[0]["device_type"] == ""


def test_import_netbox_devices_csv_maps_role_and_hardware(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "NetBox in", "site_name": "DC1"}).json()
    pid = project["id"]
    csv_body = (
        "name,role,tenant,manufacturer,device_type,serial,asset_tag,status,site,location,rack,position,face,comments\n"
        "core-sw,switch,Acme Colo,Cisco,Catalyst 9300-48P,SN-NB,AT-1,active,DC1,Hall A / Row 1,R05,42,front,core pair\n"
        "mystery-box,,,,,,active,DC1,Hall A / Row 1,R05,1,,\n"
    )
    imported = client.post(
        f"/api/projects/{pid}/import",
        headers=auth,
        files={"file": ("devices.csv", csv_body.encode(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["created"] == 2
    areas = {a["name"]: a for a in client.get(f"/api/projects/{pid}/areas", headers=auth).json()}
    assert "Hall A" in areas
    rows = {r["name"]: r for r in client.get(f"/api/projects/{pid}/rows", headers=auth).json()}
    assert "Row 1" in rows
    assert rows["Row 1"]["area_id"] == areas["Hall A"]["id"]
    racks = {r["name"]: r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json()}
    assert "R05" in racks
    devices = {d["name"]: d for d in client.get(f"/api/projects/{pid}/devices", headers=auth).json()}
    core = devices["core-sw"]
    assert core["device_type"] == "switch"
    assert core["vendor"] == "Cisco"
    assert core["model"] == "Catalyst 9300-48P"
    assert core["ru_start"] == 42
    assert core["owner"] == "Acme Colo"
    assert core["notes"] == "core pair"
    assert core["serial"] == "SN-NB"
    assert core["asset_tag"] == "AT-1"
    assert core["function"] == ""
    assert core["rack_id"] == racks["R05"]["id"]
    mystery = devices["mystery-box"]
    assert mystery["device_type"] == ""
    assert mystery["vendor"] == ""
    assert mystery["model"] == ""


def test_netbox_export_zip_roundtrip(client, auth):
    from io import BytesIO
    from zipfile import ZipFile

    src = client.post(
        "/api/projects",
        headers=auth,
        json={"name": "NB src", "site_name": "DC1", "customer": "Acme"},
    ).json()
    pid = src["id"]
    area = client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Hall A"}).json()
    row = client.post(f"/api/projects/{pid}/rows", headers=auth, json={"name": "Row 1", "area_id": area["id"]}).json()
    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "R05", "area_id": area["id"], "row_id": row["id"], "ru_height": 42},
    ).json()
    client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={
            "name": "core-sw",
            "rack_id": rack["id"],
            "vendor": "Cisco",
            "model": "Catalyst 9300-48P",
            "serial": "SN-NBX",
            "owner": "Acme Colo",
            "device_type": "switch",
            "ru_start": 42,
            "ru_end": 42,
            "notes": "core pair",
        },
    )
    client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "mystery-box", "rack_id": rack["id"], "serial": "SN-BLANK", "ru_start": 1, "ru_end": 1},
    )

    exported = client.get(f"/api/projects/{pid}/export-netbox.zip", headers=auth)
    assert exported.status_code == 200, exported.text
    assert exported.content[:2] == b"PK"
    zf = ZipFile(BytesIO(exported.content))
    names = set(zf.namelist())
    assert names >= {
        "README.txt",
        "sites.csv",
        "locations.csv",
        "racks.csv",
        "device-roles.csv",
        "manufacturers.csv",
        "devices.csv",
        "device-types.yaml",
    }
    devices_csv = zf.read("devices.csv").decode()
    header = devices_csv.splitlines()[0]
    assert header == "name,role,manufacturer,device_type,site,location,rack,position,serial,asset_tag,status,tenant,comments"
    assert "Hall A / Row 1" in devices_csv
    assert "switch" in devices_csv
    assert "unspecified" in devices_csv
    yaml_text = zf.read("device-types.yaml").decode()
    assert "manufacturer: Cisco" in yaml_text
    assert "model: Catalyst 9300-48P" in yaml_text
    locations = zf.read("locations.csv").decode()
    assert "Hall A" in locations
    assert "Hall A / Row 1" in locations

    dest = client.post("/api/projects", headers=auth, json={"name": "NB dest", "site_name": "DC1"}).json()
    dest_id = dest["id"]
    imported = client.post(
        f"/api/projects/{dest_id}/import",
        headers=auth,
        files={"file": ("site-NetBox.zip", exported.content, "application/zip")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["created"] >= 2
    got = {d["serial"]: d for d in client.get(f"/api/projects/{dest_id}/devices", headers=auth).json()}
    core = got["SN-NBX"]
    assert core["device_type"] == "switch"
    assert core["vendor"] == "Cisco"
    assert core["model"] == "Catalyst 9300-48P"
    assert core["ru_start"] == 42
    assert core["owner"] == "Acme Colo"
    blank = got["SN-BLANK"]
    assert blank["device_type"] == ""
    dest_areas = {a["name"] for a in client.get(f"/api/projects/{dest_id}/areas", headers=auth).json()}
    dest_rows = {r["name"] for r in client.get(f"/api/projects/{dest_id}/rows", headers=auth).json()}
    dest_racks = {r["name"] for r in client.get(f"/api/projects/{dest_id}/racks", headers=auth).json()}
    assert "Hall A" in dest_areas
    assert "Row 1" in dest_rows
    assert "R05" in dest_racks


def test_hierarchy_government_emss_tags_block_photos(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Gov site", "restriction_type": "government"}).json()
    pid = project["id"]
    assert project["restricted"] is True
    assert project["photography_allowed"] is False
    assert project["restriction_type"] == "government"

    area = client.post(
        f"/api/projects/{pid}/areas",
        headers=auth,
        json={"name": "Cage 7", "restriction_type": "EMSS"},
    ).json()
    assert area["restricted"] is True
    assert area["photography_allowed"] is False
    assert area["restriction_type"] == "EMSS"

    rows = client.post(
        f"/api/projects/{pid}/rows/bulk",
        headers=auth,
        json={"area_id": area["id"], "names": ["R1", "R2"], "restriction_type": "government"},
    ).json()
    created = {r["name"]: r for r in rows["created"]}
    assert created["R1"]["restricted"] is True
    assert created["R1"]["restriction_type"] == "government"
    assert created["R1"]["photography_allowed"] is False

    rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "A01", "row_id": created["R1"]["id"], "restriction_type": "EMSS"},
    ).json()
    assert rack["restricted"] is True
    assert rack["restriction_type"] == "EMSS"
    assert rack["photography_allowed"] is False

    renamed = client.patch(
        f"/api/projects/{pid}/rows/{created['R1']['id']}",
        headers=auth,
        json={"name": "Row 1"},
    ).json()
    assert renamed["name"] == "Row 1"
    assert renamed["restriction_type"] == "government"
    assert renamed["restricted"] is True

    cleared = client.patch(
        f"/api/projects/{pid}/racks/{rack['id']}",
        headers=auth,
        json={"name": rack["name"], "restricted": False, "restriction_type": "", "photography_allowed": True},
    ).json()
    assert cleared["restricted"] is False
    assert cleared["photography_allowed"] is True

    device = client.post(
        f"/api/projects/{pid}/devices",
        headers=auth,
        json={"name": "emss-box", "restricted_reason": "EMSS"},
    ).json()
    assert device["restricted"] is True
    assert device["restricted_reason"] == "EMSS"


def test_emss_tag_is_independent_per_row_and_rack(client, auth):
    project = client.post("/api/projects", headers=auth, json={"name": "Area1 site"}).json()
    pid = project["id"]
    area = client.post(f"/api/projects/{pid}/areas", headers=auth, json={"name": "Area1"}).json()
    rows = client.post(
        f"/api/projects/{pid}/rows/bulk",
        headers=auth,
        json={"area_id": area["id"], "names": ["A01", "A04"]},
    ).json()["created"]
    by_name = {r["name"]: r for r in rows}

    tagged = client.patch(
        f"/api/projects/{pid}/rows/{by_name['A01']['id']}",
        headers=auth,
        json={"name": "A01", "restriction_type": "EMSS"},
    ).json()
    listed = {r["name"]: r for r in client.get(f"/api/projects/{pid}/rows", headers=auth).json()}
    assert tagged["restriction_type"] == "EMSS"
    assert tagged["restricted"] is True
    assert tagged["photography_allowed"] is False
    assert listed["A04"]["restriction_type"] == ""
    assert listed["A04"]["restricted"] is False
    assert listed["A04"]["photography_allowed"] is True

    r1 = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "R1", "row_id": by_name["A01"]["id"]},
    ).json()
    r2 = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "R2", "row_id": by_name["A01"]["id"]},
    ).json()
    tagged_rack = client.patch(
        f"/api/projects/{pid}/racks/{r1['id']}",
        headers=auth,
        json={"name": "R1", "restriction_type": "government"},
    ).json()
    racks = {r["id"]: r for r in client.get(f"/api/projects/{pid}/racks", headers=auth).json()}
    assert tagged_rack["restriction_type"] == "government"
    assert tagged_rack["restricted"] is True
    assert racks[r2["id"]]["restriction_type"] == ""
    assert racks[r2["id"]]["restricted"] is False
    assert racks[r2["id"]]["photography_allowed"] is True

    a04_rack = client.post(
        f"/api/projects/{pid}/racks",
        headers=auth,
        json={"name": "R4", "row_id": by_name["A04"]["id"]},
    ).json()
    assert a04_rack["restricted"] is False
    assert a04_rack["restriction_type"] == ""








