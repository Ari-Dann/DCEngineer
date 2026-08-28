"""Office / Visio export: hierarchy-preserving ZIP with .vsdx, Excel, pictures, and SVG elevations."""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from app.models import AisleRow, Area, Attachment, Device, Project, Rack
from app.rbi_export import HEADER_FILL, HEADER_FONT, _autosize, _header, rack_svg
from app.storage import get_storage

V_NS = "http://schemas.microsoft.com/office/visio/2012/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

IMAGE_EXTS = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/gif": ".gif",
    "image/bmp": ".bmp",
    "image/webp": ".webp",
}
VSDX_IMAGE_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/gif", "image/bmp"}

KIND_FILL = {
    "Area": "#1B3A4B",
    "Row": "#2A5470",
    "Rack": "#3D9CF0",
    "Device": "#3DBF8C",
    "Unlocated": "#8B9BB0",
}
DEVICE_FILL = {
    "server": "#3D9CF0",
    "switch": "#3DBF8C",
    "firewall": "#E87A4C",
    "router": "#5EC8E8",
    "storage": "#9B7DFF",
    "pdu": "#E8A317",
    "ups": "#E85D4C",
}


def _safe(name: str, fallback: str = "unnamed") -> str:
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", (name or "").strip())
    text = re.sub(r"\s+", " ", text).strip(" ._") or fallback
    return text[:80]


def _unique(name: str, used: set[str], limit: int = 48) -> str:
    base = (name or "Page")[:limit]
    candidate = base
    n = 2
    while candidate.lower() in used:
        suffix = f" {n}"
        candidate = f"{base[: limit - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate.lower())
    return candidate


def _xml(text: str) -> str:
    return escape(text or "", {'"': "&quot;", "'": "&apos;"})


@dataclass
class Picture:
    zip_path: str
    data: bytes
    content_type: str
    filename: str
    entity_type: str
    entity_id: int
    label: str


@dataclass
class Node:
    visio_id: str
    name: str
    title: str
    manager_id: str
    extra: dict[str, str] = field(default_factory=dict)
    pictures: list[Picture] = field(default_factory=list)
    page_name: str = ""


@dataclass
class Layout:
    project: Project
    areas: list[Area]
    rows: list[AisleRow]
    racks: list[Rack]
    devices: list[Device]
    area_by_id: dict[int, Area]
    row_by_id: dict[int, AisleRow]
    rack_by_id: dict[int, Rack]
    devices_by_rack: dict[int | None, list[Device]]
    pictures: list[Picture]
    pictures_by_key: dict[tuple[str, int], list[Picture]]
    nodes: list[Node]
    skipped_restricted: int = 0


def _ext_for(att: Attachment) -> str:
    ctype = (att.content_type or "").lower()
    if ctype in IMAGE_EXTS:
        return IMAGE_EXTS[ctype]
    name = (att.filename or "").lower()
    for ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"):
        if name.endswith(ext):
            return ".jpg" if ext == ".jpeg" else ext
    return ".bin"


def _folder_for(layout: Layout, entity_type: str, entity_id: int) -> str:
    if entity_type == "project":
        return "Pictures/_project"
    if entity_type == "area":
        area = layout.area_by_id.get(entity_id)
        return f"Pictures/{_safe(area.name if area else f'area-{entity_id}')}/_area"
    if entity_type == "rack":
        rack = layout.rack_by_id.get(entity_id)
        if not rack:
            return f"Pictures/_racks/{entity_id}"
        row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
        area = layout.area_by_id.get(rack.area_id) if rack.area_id else None
        if not area and row and row.area_id:
            area = layout.area_by_id.get(row.area_id)
        parts = ["Pictures"]
        parts.append(_safe(area.name) if area else "_unassigned")
        parts.append(_safe(row.name if row else rack.row_label or "_no-row"))
        parts.append(_safe(rack.name))
        return "/".join(parts)
    if entity_type == "device":
        device = next((d for d in layout.devices if d.id == entity_id), None)
        if not device:
            return f"Pictures/_devices/{entity_id}"
        rack = layout.rack_by_id.get(device.rack_id) if device.rack_id else None
        if rack:
            return f"{_folder_for(layout, 'rack', rack.id)}/{_safe(device.name)}"
        return f"Pictures/_unlocated/{_safe(device.name)}"
    return f"Pictures/{entity_type}/{entity_id}"


def collect_layout(db: Session, project: Project) -> Layout:
    areas = db.query(Area).filter(Area.project_id == project.id).order_by(Area.name).all()
    rows = db.query(AisleRow).filter(AisleRow.project_id == project.id).order_by(AisleRow.name).all()
    racks = db.query(Rack).filter(Rack.project_id == project.id).order_by(Rack.name).all()
    devices = db.query(Device).filter(Device.project_id == project.id).order_by(Device.name).all()
    layout = Layout(
        project=project,
        areas=areas,
        rows=rows,
        racks=racks,
        devices=devices,
        area_by_id={a.id: a for a in areas},
        row_by_id={r.id: r for r in rows},
        rack_by_id={r.id: r for r in racks},
        devices_by_rack={},
        pictures=[],
        pictures_by_key={},
        nodes=[],
    )
    for device in devices:
        layout.devices_by_rack.setdefault(device.rack_id, []).append(device)

    entity_ids: dict[str, list[int]] = {
        "project": [project.id],
        "area": [a.id for a in areas],
        "rack": [r.id for r in racks],
        "device": [d.id for d in devices],
    }
    attachments: list[Attachment] = []
    for entity_type, ids in entity_ids.items():
        if not ids:
            continue
        attachments.extend(
            db.query(Attachment)
            .filter(Attachment.entity_type == entity_type, Attachment.entity_id.in_(ids))
            .order_by(Attachment.created_at.asc())
            .all()
        )

    used_paths: set[str] = set()
    storage = get_storage()
    skipped = 0
    for att in attachments:
        if att.photography_restricted:
            skipped += 1
            continue
        try:
            data = storage.get(att.storage_key)
        except Exception:
            continue
        folder = _folder_for(layout, att.entity_type, att.entity_id)
        base = _safe(att.filename.rsplit(".", 1)[0] if att.filename else f"photo-{att.id}")
        ext = _ext_for(att)
        zip_path = f"{folder}/{base}{ext}"
        n = 2
        while zip_path.lower() in used_paths:
            zip_path = f"{folder}/{base}-{n}{ext}"
            n += 1
        used_paths.add(zip_path.lower())
        label = att.filename or base
        pic = Picture(
            zip_path=zip_path,
            data=data,
            content_type=att.content_type or "application/octet-stream",
            filename=att.filename or f"{base}{ext}",
            entity_type=att.entity_type,
            entity_id=att.entity_id,
            label=label,
        )
        layout.pictures.append(pic)
        layout.pictures_by_key.setdefault((att.entity_type, att.entity_id), []).append(pic)
    layout.skipped_restricted = skipped

    nodes: list[Node] = [
        Node(
            visio_id="SITE",
            name=project.site_name or project.name,
            title="Site",
            manager_id="",
            extra={
                "Area": "",
                "Row": "",
                "Rack": "",
                "RU": "",
                "Serial": "",
                "Owner": project.customer or "",
                "Vendor": "",
                "Model": "",
                "Picture": "; ".join(p.zip_path for p in layout.pictures_by_key.get(("project", project.id), [])),
                "Notes": project.in_scope_summary or "",
            },
            pictures=layout.pictures_by_key.get(("project", project.id), []),
            page_name="Overview",
        )
    ]
    for area in areas:
        pics = layout.pictures_by_key.get(("area", area.id), [])
        nodes.append(
            Node(
                visio_id=f"AREA-{area.id}",
                name=area.name,
                title="Area",
                manager_id="SITE",
                extra={
                    "Area": area.name,
                    "Row": "",
                    "Rack": "",
                    "RU": "",
                    "Serial": "",
                    "Owner": "",
                    "Vendor": "",
                    "Model": "",
                    "Picture": "; ".join(p.zip_path for p in pics),
                    "Notes": area.description or area.restriction_type or "",
                },
                pictures=pics,
            )
        )
    for row in rows:
        area = layout.area_by_id.get(row.area_id) if row.area_id else None
        nodes.append(
            Node(
                visio_id=f"ROW-{row.id}",
                name=row.name,
                title="Row",
                manager_id=f"AREA-{area.id}" if area else "SITE",
                extra={
                    "Area": area.name if area else "",
                    "Row": row.name,
                    "Rack": "",
                    "RU": "",
                    "Serial": "",
                    "Owner": "",
                    "Vendor": "",
                    "Model": "",
                    "Picture": "",
                    "Notes": row.notes or "",
                },
            )
        )
    for rack in racks:
        row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
        area = layout.area_by_id.get(rack.area_id) if rack.area_id else None
        if not area and row and row.area_id:
            area = layout.area_by_id.get(row.area_id)
        parent = f"ROW-{row.id}" if row else (f"AREA-{area.id}" if area else "SITE")
        pics = layout.pictures_by_key.get(("rack", rack.id), [])
        nodes.append(
            Node(
                visio_id=f"RACK-{rack.id}",
                name=rack.name,
                title="Rack",
                manager_id=parent,
                extra={
                    "Area": area.name if area else "",
                    "Row": (row.name if row else "") or rack.row_label,
                    "Rack": rack.name,
                    "RU": str(rack.ru_height or ""),
                    "Serial": "",
                    "Owner": "",
                    "Vendor": "",
                    "Model": "",
                    "Picture": "; ".join(p.zip_path for p in pics),
                    "Notes": rack.notes or "",
                },
                pictures=pics,
            )
        )
    unlocated = layout.devices_by_rack.get(None, [])
    if unlocated:
        nodes.append(
            Node(
                visio_id="UNLOCATED",
                name="Unlocated devices",
                title="Unlocated",
                manager_id="SITE",
                extra={"Area": "", "Row": "", "Rack": "", "RU": "", "Serial": "", "Owner": "", "Vendor": "", "Model": "", "Picture": "", "Notes": ""},
            )
        )
    for device in devices:
        rack = layout.rack_by_id.get(device.rack_id) if device.rack_id else None
        parent = f"RACK-{rack.id}" if rack else "UNLOCATED"
        ru = ""
        if device.ru_start:
            ru = str(device.ru_start) if not device.ru_end or device.ru_end == device.ru_start else f"{device.ru_start}-{device.ru_end}"
        pics = layout.pictures_by_key.get(("device", device.id), [])
        area_name = ""
        row_name = ""
        if rack:
            row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
            area = layout.area_by_id.get(rack.area_id) if rack.area_id else None
            if not area and row and row.area_id:
                area = layout.area_by_id.get(row.area_id)
            area_name = area.name if area else ""
            row_name = (row.name if row else "") or rack.row_label
        nodes.append(
            Node(
                visio_id=f"DEV-{device.id}",
                name=device.name,
                title="Device",
                manager_id=parent,
                extra={
                    "Area": area_name,
                    "Row": row_name,
                    "Rack": rack.name if rack else "",
                    "RU": ru,
                    "Serial": device.serial or "",
                    "Owner": getattr(device, "owner", "") or "",
                    "Vendor": device.vendor or "",
                    "Model": device.model or "",
                    "Picture": "; ".join(p.zip_path for p in pics),
                    "Notes": device.notes or "",
                },
                pictures=pics,
            )
        )
    layout.nodes = nodes
    return layout


def _readme(layout: Layout, stem: str) -> str:
    return f"""DCEngineer — Visio / Microsoft Office export
Site: {layout.project.site_name or layout.project.name}
Exported: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}

This folder is meant for a Windows VM on Azure running Microsoft 365 / Office and Visio.

Open in Visio
  1. Keep this folder together (do not separate Pictures from the drawings).
  2. Double-click "{stem}.vsdx".
  3. Pages:
       Overview     — areas on the site
       each Area    — rows and racks
       each Rack    — RU elevation plus associated photos
  4. Click an area or rack box to jump to its page (hyperlink).
  5. Extra photos live under Pictures\\Area\\Row\\Rack\\Device\\ so you can
     Insert → Pictures in Visio, Word, or PowerPoint.

Build a live Visio org chart from Excel (Data Visualizer)
  1. Open "{stem} - Data Visualizer.xlsx".
  2. Select the VisioHierarchy table.
  3. Insert → Add-ins → Visio Data Visualizer → Organization Chart
     (Visio Plan 2 / Microsoft 365 Apps with the add-in).
  4. Map columns:
       Employee ID  = ID
       Name         = Name
       Title        = Title
       Reports to   = Manager ID
  5. Picture Path points at files in this folder if you want to attach photos.

SVG elevations
  Elevations\\*.svg can be Insert → Pictures in Visio or PowerPoint. Visio can
  convert SVG to editable Office shapes (right-click → Convert to Shape).

Restricted photography
  {layout.skipped_restricted} photo(s) marked photography-restricted were omitted.
"""


def build_visualizer_workbook(layout: Layout) -> bytes:
    wb = Workbook()
    how = wb.active
    how.title = "How to open"
    how["A1"] = "Visio Data Visualizer — datacenter hierarchy"
    how["A1"].font = Font(size=16, bold=True, color="1B3A4B")
    how["A3"] = (
        "The VisioHierarchy table is an organization chart of Site → Area → Row → Rack → Device. "
        "In Excel (Microsoft 365, Azure VM): Insert → Visio Data Visualizer → Organization Chart. "
        "Map ID, Name, Title, and Manager ID. Picture Path is relative to this unzipped folder."
    )
    how["A3"].alignment = Alignment(wrap_text=True)
    how.merge_cells("A3:F6")
    how.column_dimensions["A"].width = 28
    how.row_dimensions[3].height = 80

    vis = wb.create_sheet("VisioHierarchy")
    headers = [
        "ID",
        "Name",
        "Title",
        "Manager ID",
        "Area",
        "Row",
        "Rack",
        "RU",
        "Serial",
        "Owner",
        "Vendor",
        "Model",
        "Picture Path",
        "Notes",
    ]
    for col, title in enumerate(headers, 1):
        cell = vis.cell(1, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for node in layout.nodes:
        vis.append(
            [
                node.visio_id,
                node.name,
                node.title,
                node.manager_id,
                node.extra.get("Area", ""),
                node.extra.get("Row", ""),
                node.extra.get("Rack", ""),
                node.extra.get("RU", ""),
                node.extra.get("Serial", ""),
                node.extra.get("Owner", ""),
                node.extra.get("Vendor", ""),
                node.extra.get("Model", ""),
                node.extra.get("Picture", ""),
                node.extra.get("Notes", ""),
            ]
        )
    last = vis.max_row
    if last >= 2:
        table = Table(displayName="VisioHierarchy", ref=f"A1:{get_column_letter(len(headers))}{last}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        vis.add_table(table)
    vis.freeze_panes = "A2"
    _autosize(vis)

    pics = wb.create_sheet("Pictures")
    _header(pics, ["Hierarchy", "Kind", "File", "Path"])
    for pic in layout.pictures:
        path = pic.zip_path
        kind = pic.entity_type
        pics.append([pic.label, kind, pic.filename, path])
        cell = pics.cell(pics.max_row, 4)
        cell.hyperlink = path
        cell.style = "Hyperlink"
    _autosize(pics)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(n: str, v: str, extra: str = "") -> str:
    return f'<Cell N="{n}" V="{_xml(str(v))}"{extra}/>'


def _rect_geometry() -> str:
    return """
      <Section N="Geometry" IX="0">
        <Cell N="NoFill" V="0"/>
        <Cell N="NoLine" V="0"/>
        <Cell N="NoShow" V="0"/>
        <Cell N="NoSnap" V="0"/>
        <Row T="RelMoveTo" IX="1"><Cell N="X" V="0"/><Cell N="Y" V="0"/></Row>
        <Row T="RelLineTo" IX="2"><Cell N="X" V="1"/><Cell N="Y" V="0"/></Row>
        <Row T="RelLineTo" IX="3"><Cell N="X" V="1"/><Cell N="Y" V="1"/></Row>
        <Row T="RelLineTo" IX="4"><Cell N="X" V="0"/><Cell N="Y" V="1"/></Row>
        <Row T="RelLineTo" IX="5"><Cell N="X" V="0"/><Cell N="Y" V="0"/></Row>
      </Section>"""


def _shape_xml(
    shape_id: int,
    *,
    pin_x: float,
    pin_y: float,
    width: float,
    height: float,
    text: str,
    fill: str,
    font_color: str = "#FFFFFF",
    font_size_in: float = 0.12,
    hyperlink_page: str = "",
    shape_type: str = "Shape",
    extra_cells: str = "",
    extra_body: str = "",
) -> str:
    hyper = ""
    if hyperlink_page:
        hyper = f"""
      <Section N="Hyperlink" IX="0">
        {_cell("Description", hyperlink_page)}
        {_cell("Address", "")}
        {_cell("SubAddress", hyperlink_page)}
        {_cell("NewWindow", "0")}
        {_cell("Default", "")}
        {_cell("ExtraInfo", "")}
        {_cell("Frame", "")}
        {_cell("SortKey", "")}
        {_cell("Invisible", "0")}
      </Section>"""
    text_xml = f"<Text>{_xml(text)}</Text>" if text else ""
    return f"""
    <Shape ID="{shape_id}" Type="{shape_type}" LineStyle="0" FillStyle="0" TextStyle="0">
      {_cell("PinX", f"{pin_x:.4f}")}
      {_cell("PinY", f"{pin_y:.4f}")}
      {_cell("Width", f"{width:.4f}")}
      {_cell("Height", f"{height:.4f}")}
      {_cell("LocPinX", f"{width / 2:.4f}", ' F="Width*0.5"')}
      {_cell("LocPinY", f"{height / 2:.4f}", ' F="Height*0.5"')}
      {_cell("Angle", "0")}
      {_cell("FlipX", "0")}
      {_cell("FlipY", "0")}
      {_cell("ResizeMode", "0")}
      {_cell("FillForegnd", fill)}
      {_cell("FillBkgnd", "#FFFFFF")}
      {_cell("FillPattern", "1")}
      {_cell("LineColor", "#1B3A4B")}
      {_cell("LinePattern", "1")}
      {_cell("LineWeight", "0.01")}
      {_cell("LeftMargin", "0.04")}
      {_cell("RightMargin", "0.04")}
      {_cell("TopMargin", "0.02")}
      {_cell("BottomMargin", "0.02")}
      {_cell("VerticalAlign", "1")}
      {extra_cells}
      <Section N="Character">
        {_cell("Color", font_color)}
        {_cell("Size", f"{font_size_in:.4f}")}
        {_cell("Font", "Calibri")}
        {_cell("Style", "0")}
      </Section>
      {_rect_geometry()}
      {hyper}
      {extra_body}
      {text_xml}
    </Shape>"""


def _foreign_image_shape(shape_id: int, rel_id: str, pin_x: float, pin_y: float, width: float, height: float, caption: str) -> str:
    extra_cells = (
        f'{_cell("ImgOffsetX", "0")}{_cell("ImgOffsetY", "0")}'
        f'{_cell("ImgWidth", f"{width:.4f}", " F=&quot;Width*1&quot;")}{_cell("ImgHeight", f"{height:.4f}", " F=&quot;Height*1&quot;")}'
    )
    extra_body = f"""
      <ForeignData ForeignType="Bitmap">
        <Rel r:id="{rel_id}"/>
      </ForeignData>"""
    return _shape_xml(
        shape_id,
        pin_x=pin_x,
        pin_y=pin_y,
        width=width,
        height=height,
        text=caption,
        fill="#FFFFFF",
        font_color="#1B3A4B",
        font_size_in=0.08,
        shape_type="Foreign",
        extra_cells=extra_cells,
        extra_body=extra_body,
    )


@dataclass
class VsdxPage:
    page_id: int
    name: str
    filename: str
    rel_id: str
    width: float = 17.0
    height: float = 11.0
    shapes: list[str] = field(default_factory=list)
    images: list[tuple[str, bytes, str]] = field(default_factory=list)  # rel_id, data, ext


class VsdxBuilder:
    def __init__(self, title: str):
        self.title = title
        self.pages: list[VsdxPage] = []
        self._used_names: set[str] = set()
        self._image_n = 0

    def add_page(self, name: str, width: float = 17.0, height: float = 11.0) -> VsdxPage:
        page_id = len(self.pages)
        filename = f"page{page_id + 1}.xml"
        page = VsdxPage(
            page_id=page_id,
            name=_unique(name, self._used_names),
            filename=filename,
            rel_id=f"rId{page_id + 1}",
            width=width,
            height=height,
        )
        self.pages.append(page)
        return page

    def add_image(self, page: VsdxPage, data: bytes, content_type: str) -> str:
        self._image_n += 1
        ext = IMAGE_EXTS.get((content_type or "").lower(), ".jpg")
        if ext == ".webp":
            ext = ".jpg"
        rel_id = f"rIdImg{self._image_n}"
        page.images.append((rel_id, data, ext))
        return rel_id

    def dumps(self) -> bytes:
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            defaults = [
                ("rels", "application/vnd.openxmlformats-package.relationships+xml"),
                ("xml", "application/xml"),
                ("jpg", "image/jpeg"),
                ("jpeg", "image/jpeg"),
                ("png", "image/png"),
                ("gif", "image/gif"),
                ("bmp", "image/bmp"),
            ]
            overrides = [
                ("/visio/document.xml", "application/vnd.ms-visio.drawing.main+xml"),
                ("/visio/pages/pages.xml", "application/vnd.ms-visio.pages+xml"),
                ("/visio/windows.xml", "application/vnd.ms-visio.windows+xml"),
                ("/docProps/core.xml", "application/vnd.openxmlformats-package.core-properties+xml"),
                ("/docProps/app.xml", "application/vnd.openxmlformats-officedocument.extended-properties+xml"),
            ]
            for page in self.pages:
                overrides.append((f"/visio/pages/{page.filename}", "application/vnd.ms-visio.page+xml"))
            ct = [f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>', f'<Types xmlns="{CT_NS}">']
            for ext, mime in defaults:
                ct.append(f'<Default Extension="{ext}" ContentType="{mime}"/>')
            for part, mime in overrides:
                ct.append(f'<Override PartName="{part}" ContentType="{mime}"/>')
            ct.append("</Types>")
            zf.writestr("[Content_Types].xml", "\n".join(ct))

            zf.writestr(
                "_rels/.rels",
                f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{PKG_NS}">
  <Relationship Id="rId1" Type="http://schemas.microsoft.com/visio/2010/relationships/document" Target="visio/document.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>""",
            )
            now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            zf.writestr(
                "docProps/core.xml",
                f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>{_xml(self.title)}</dc:title>
  <dc:creator>DCEngineer</dc:creator>
  <cp:lastModifiedBy>DCEngineer</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>
</cp:coreProperties>""",
            )
            zf.writestr(
                "docProps/app.xml",
                f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>DCEngineer</Application>
  <Pages>{len(self.pages)}</Pages>
</Properties>""",
            )
            zf.writestr(
                "visio/_rels/document.xml.rels",
                f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{PKG_NS}">
  <Relationship Id="rId1" Type="http://schemas.microsoft.com/visio/2010/relationships/pages" Target="pages/pages.xml"/>
  <Relationship Id="rId2" Type="http://schemas.microsoft.com/visio/2010/relationships/windows" Target="windows.xml"/>
</Relationships>""",
            )
            zf.writestr("visio/document.xml", _document_xml())
            first = self.pages[0].page_id if self.pages else 0
            zf.writestr(
                "visio/windows.xml",
                f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Windows xmlns="{V_NS}" ClientWidth="1600" ClientHeight="900">
  <Window ID="0" WindowType="Drawing" WindowState="1073741824" WindowLeft="0" WindowTop="0" WindowWidth="1600" WindowHeight="900" ContainerType="Page" Page="{first}" ViewScale="-1" ViewCenterX="{self.pages[0].width / 2 if self.pages else 8.5}" ViewCenterY="{self.pages[0].height / 2 if self.pages else 5.5}">
    <ShowRulers>1</ShowRulers>
    <ShowGrid>1</ShowGrid>
    <ShowPageBreaks>0</ShowPageBreaks>
    <ShowGuides>1</ShowGuides>
    <ShowConnectionPoints>1</ShowConnectionPoints>
  </Window>
</Windows>""",
            )
            pages_rels = [f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>', f'<Relationships xmlns="{PKG_NS}">']
            pages_xml = [
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                f'<Pages xmlns="{V_NS}" xmlns:r="{R_NS}" xml:space="preserve">',
            ]
            media_n = 0
            for page in self.pages:
                pages_rels.append(
                    f'<Relationship Id="{page.rel_id}" Type="http://schemas.microsoft.com/visio/2010/relationships/page" Target="{page.filename}"/>'
                )
                pages_xml.append(
                    f"""  <Page ID="{page.page_id}" Name="{_xml(page.name)}" NameU="{_xml(page.name)}" ViewScale="0.7" ViewCenterX="{page.width / 2:.4f}" ViewCenterY="{page.height / 2:.4f}">
    <PageSheet LineStyle="0" FillStyle="0" TextStyle="0">
      {_cell("PageWidth", f"{page.width}")}
      {_cell("PageHeight", f"{page.height}")}
      {_cell("PageScale", "1")}
      {_cell("DrawingScale", "1")}
      {_cell("DrawingSizeType", "1")}
      {_cell("DrawingScaleType", "0")}
      {_cell("DrawingResizeType", "0")}
    </PageSheet>
    <Rel r:id="{page.rel_id}"/>
  </Page>"""
                )
                shapes = "\n".join(page.shapes) if page.shapes else ""
                zf.writestr(
                    f"visio/pages/{page.filename}",
                    f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<PageContents xmlns="{V_NS}" xmlns:r="{R_NS}" xml:space="preserve">
  <Shapes>
{shapes}
  </Shapes>
</PageContents>""",
                )
                if page.images:
                    rels = [f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>', f'<Relationships xmlns="{PKG_NS}">']
                    for rel_id, data, ext in page.images:
                        media_n += 1
                        media_name = f"image{media_n}{ext}"
                        zf.writestr(f"visio/media/{media_name}", data)
                        rels.append(
                            f'<Relationship Id="{rel_id}" Type="http://schemas.microsoft.com/visio/2010/relationships/image" Target="../media/{media_name}"/>'
                        )
                    rels.append("</Relationships>")
                    zf.writestr(f"visio/pages/_rels/{page.filename}.rels", "\n".join(rels))
            pages_rels.append("</Relationships>")
            pages_xml.append("</Pages>")
            zf.writestr("visio/pages/_rels/pages.xml.rels", "\n".join(pages_rels))
            zf.writestr("visio/pages/pages.xml", "\n".join(pages_xml))
        return buf.getvalue()


def _document_xml() -> str:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<VisioDocument xmlns="{V_NS}" xmlns:r="{R_NS}" xml:space="preserve">
  <DocumentSettings TopPage="0" DefaultTextStyle="0" DefaultLineStyle="0" DefaultFillStyle="0">
    <GlueSettings>9</GlueSettings>
    <SnapSettings>65887</SnapSettings>
    <DynamicGridEnabled>1</DynamicGridEnabled>
  </DocumentSettings>
  <Colors>
    <ColorEntry IX="0" RGB="#000000"/>
    <ColorEntry IX="1" RGB="#FFFFFF"/>
  </Colors>
  <FaceNames>
    <FaceName ID="0" NameU="Calibri" UnicodeRanges="-1 -1 0 0 0" CharSets="536870911 0" Panose="2 15 5 2 2 2 4 3 2 4" Flags="325"/>
  </FaceNames>
  <StyleSheets>
    <StyleSheet ID="0" NameU="No Style" Name="No Style">
      <Cell N="EnableLineProps" V="1"/>
      <Cell N="EnableFillProps" V="1"/>
      <Cell N="EnableTextProps" V="1"/>
      <Cell N="LineWeight" V="0.01"/>
      <Cell N="LineColor" V="#1B3A4B"/>
      <Cell N="LinePattern" V="1"/>
      <Cell N="FillForegnd" V="#FFFFFF"/>
      <Cell N="FillBkgnd" V="#FFFFFF"/>
      <Cell N="FillPattern" V="1"/>
      <Section N="Character">
        <Cell N="Font" V="Calibri"/>
        <Cell N="Color" V="#1B3A4B"/>
        <Cell N="Style" V="0"/>
        <Cell N="Size" V="0.1389"/>
      </Section>
    </StyleSheet>
  </StyleSheets>
</VisioDocument>"""


def _grid_positions(count: int, page_w: float, page_h: float, box_w: float, box_h: float, top: float, margin: float = 0.4):
    if count <= 0:
        return []
    usable = page_w - 2 * margin
    cols = max(1, int(usable // (box_w + 0.25)))
    gap_x = (usable - cols * box_w) / max(cols, 1)
    positions = []
    for i in range(count):
        col = i % cols
        row = i // cols
        x = margin + box_w / 2 + col * (box_w + max(gap_x, 0.2))
        y = top - box_h / 2 - row * (box_h + 0.25)
        if y - box_h / 2 < 0.3:
            y = 0.3 + box_h / 2
        positions.append((x, y))
    return positions


def build_vsdx(layout: Layout) -> bytes:
    title = layout.project.site_name or layout.project.name
    builder = VsdxBuilder(title)
    sid = 1

    overview = builder.add_page("Overview")
    overview.shapes.append(
        _shape_xml(
            sid,
            pin_x=8.5,
            pin_y=10.4,
            width=16.2,
            height=0.7,
            text=f"{title}  ·  {layout.project.customer or layout.project.name}",
            fill="#1B3A4B",
            font_size_in=0.2,
        )
    )
    sid += 1
    area_pages: dict[int, str] = {}
    for area in layout.areas:
        area_pages[area.id] = f"Area {area.name}"[:48]
    positions = _grid_positions(len(layout.areas), 17, 11, 3.6, 1.4, 9.4)
    for area, (x, y) in zip(layout.areas, positions):
        row_count = sum(1 for r in layout.rows if r.area_id == area.id)
        rack_count = sum(
            1
            for r in layout.racks
            if r.area_id == area.id or (layout.row_by_id.get(r.row_id).area_id == area.id if r.row_id and layout.row_by_id.get(r.row_id) else False)
        )
        overview.shapes.append(
            _shape_xml(
                sid,
                pin_x=x,
                pin_y=y,
                width=3.6,
                height=1.4,
                text=f"{area.name}\n{row_count} rows · {rack_count} racks",
                fill=KIND_FILL["Area"],
                hyperlink_page=area_pages.get(area.id, ""),
            )
        )
        sid += 1
        for pic in layout.pictures_by_key.get(("area", area.id), [])[:1]:
            if (pic.content_type or "").lower() not in VSDX_IMAGE_TYPES:
                continue
            rel = builder.add_image(overview, pic.data, pic.content_type)
            overview.shapes.append(_foreign_image_shape(sid, rel, x, y - 1.3, 1.6, 1.1, ""))
            sid += 1

    unlocated = layout.devices_by_rack.get(None, [])
    if unlocated:
        overview.shapes.append(
            _shape_xml(
                sid,
                pin_x=2.0,
                pin_y=0.7,
                width=3.6,
                height=0.8,
                text=f"Unlocated devices ({len(unlocated)})",
                fill=KIND_FILL["Unlocated"],
                hyperlink_page="Unlocated",
            )
        )
        sid += 1

    rack_page_names: dict[int, str] = {}
    used_rack_pages: set[str] = set()
    for rack in layout.racks:
        row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
        label = f"Rack {rack.name}"
        if row:
            label = f"Rack {row.name}-{rack.name}"
        rack_page_names[rack.id] = _unique(label, used_rack_pages)

    for area in layout.areas:
        page = builder.add_page(area_pages[area.id])
        page.shapes.append(
            _shape_xml(
                sid,
                pin_x=8.5,
                pin_y=10.4,
                width=16.2,
                height=0.7,
                text=f"Area {area.name}  ·  click a rack for elevation",
                fill="#1B3A4B",
                hyperlink_page="Overview",
            )
        )
        sid += 1
        area_rows = [r for r in layout.rows if r.area_id == area.id]
        loose_racks = [
            r
            for r in layout.racks
            if (r.area_id == area.id and not r.row_id)
            or (r.area_id == area.id and r.row_id and layout.row_by_id.get(r.row_id) is None)
        ]
        bands = [(row.name, [rk for rk in layout.racks if rk.row_id == row.id]) for row in area_rows]
        if loose_racks:
            bands.append(("Unassigned row", loose_racks))
        y = 9.5
        for row_name, racks in bands:
            page.shapes.append(
                _shape_xml(
                    sid,
                    pin_x=8.5,
                    pin_y=y,
                    width=16.2,
                    height=0.35,
                    text=f"Row {row_name}",
                    fill=KIND_FILL["Row"],
                    font_size_in=0.11,
                )
            )
            sid += 1
            y -= 0.55
            if not racks:
                y -= 0.2
                continue
            box_w = min(2.4, 15.4 / max(len(racks), 1))
            for i, rack in enumerate(racks):
                x = 0.8 + box_w / 2 + i * (box_w + 0.15)
                n_dev = len(layout.devices_by_rack.get(rack.id, []))
                page.shapes.append(
                    _shape_xml(
                        sid,
                        pin_x=x,
                        pin_y=y - 0.45,
                        width=box_w,
                        height=0.9,
                        text=f"{rack.name}\n{rack.ru_height}U · {n_dev} devices",
                        fill=KIND_FILL["Rack"],
                        font_color="#0B0F14",
                        hyperlink_page=rack_page_names.get(rack.id, ""),
                    )
                )
                sid += 1
            y -= 1.25
            if y < 1.2:
                break

    for rack in layout.racks:
        page = builder.add_page(rack_page_names[rack.id], width=11.0, height=17.0)
        row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
        area = layout.area_by_id.get(rack.area_id) if rack.area_id else None
        if not area and row and row.area_id:
            area = layout.area_by_id.get(row.area_id)
        path = " / ".join(p for p in (area.name if area else "", row.name if row else rack.row_label, rack.name) if p)
        page.shapes.append(
            _shape_xml(
                sid,
                pin_x=5.5,
                pin_y=16.4,
                width=10.2,
                height=0.7,
                text=f"{path}  ·  {rack.ru_height}U",
                fill="#1B3A4B",
                hyperlink_page=area_pages.get(area.id, "Overview") if area else "Overview",
            )
        )
        sid += 1
        ru = rack.ru_height or 42
        slot_h = min(0.32, 14.8 / max(ru, 1))
        elev_w = 4.6
        top_y = 15.6
        occupied: dict[int, Device] = {}
        for device in layout.devices_by_rack.get(rack.id, []):
            if device.ru_start is None:
                continue
            start, end = int(device.ru_start), int(device.ru_end or device.ru_start)
            for u in range(min(start, end), max(start, end) + 1):
                occupied[u] = device
        drawn: set[int] = set()
        for u in range(ru, 0, -1):
            y = top_y - (ru - u + 0.5) * slot_h
            page.shapes.append(
                _shape_xml(
                    sid,
                    pin_x=0.45,
                    pin_y=y,
                    width=0.4,
                    height=slot_h * 0.92,
                    text=str(u),
                    fill="#121820",
                    font_size_in=0.07,
                )
            )
            sid += 1
            device = occupied.get(u)
            if device and device.id not in drawn and (device.ru_end or device.ru_start) == u:
                start, end = int(device.ru_start), int(device.ru_end or device.ru_start)
                span = abs(end - start) + 1
                top_u = max(start, end)
                cy = top_y - (ru - top_u + span / 2) * slot_h
                owner = getattr(device, "owner", "") or ""
                label = f"{device.name}  {device.vendor} {device.model}".strip()
                if owner:
                    label += f"  ·  {owner}"
                page.shapes.append(
                    _shape_xml(
                        sid,
                        pin_x=2.9,
                        pin_y=cy,
                        width=elev_w,
                        height=span * slot_h * 0.92,
                        text=label,
                        fill=DEVICE_FILL.get(device.device_type, "#8B9BB0"),
                        font_color="#0B0F14",
                        font_size_in=0.08,
                    )
                )
                sid += 1
                drawn.add(device.id)
            elif not device:
                page.shapes.append(
                    _shape_xml(
                        sid,
                        pin_x=2.9,
                        pin_y=y,
                        width=elev_w,
                        height=slot_h * 0.92,
                        text="",
                        fill="#E8EEF6",
                        font_color="#8B9BB0",
                    )
                )
                sid += 1
        photo_x = 8.2
        photo_y = 15.2
        photos: list[tuple[str, Picture]] = []
        for pic in layout.pictures_by_key.get(("rack", rack.id), []):
            photos.append((f"Rack · {pic.label}", pic))
        for device in layout.devices_by_rack.get(rack.id, []):
            for pic in layout.pictures_by_key.get(("device", device.id), []):
                photos.append((f"{device.name} · {pic.label}", pic))
        for caption, pic in photos[:8]:
            if (pic.content_type or "").lower() not in VSDX_IMAGE_TYPES:
                continue
            rel = builder.add_image(page, pic.data, pic.content_type)
            page.shapes.append(_foreign_image_shape(sid, rel, photo_x, photo_y, 2.4, 1.7, caption[:40]))
            sid += 1
            photo_y -= 2.05
            if photo_y < 1.2:
                photo_x += 2.6
                photo_y = 15.2

    if unlocated:
        page = builder.add_page("Unlocated")
        page.shapes.append(
            _shape_xml(
                sid,
                pin_x=8.5,
                pin_y=10.4,
                width=16.2,
                height=0.7,
                text="Unlocated devices — assign a rack later",
                fill="#1B3A4B",
                hyperlink_page="Overview",
            )
        )
        sid += 1
        positions = _grid_positions(len(unlocated), 17, 11, 3.4, 1.1, 9.4)
        for device, (x, y) in zip(unlocated, positions):
            page.shapes.append(
                _shape_xml(
                    sid,
                    pin_x=x,
                    pin_y=y,
                    width=3.4,
                    height=1.1,
                    text=f"{device.name}\n{device.vendor} {device.model}".strip(),
                    fill=DEVICE_FILL.get(device.device_type, KIND_FILL["Device"]),
                    font_color="#0B0F14",
                )
            )
            sid += 1

    if not builder.pages:
        builder.add_page("Overview")
    return builder.dumps()


def elevation_zip_path(layout: Layout, rack: Rack) -> str:
    row = layout.row_by_id.get(rack.row_id) if rack.row_id else None
    area = layout.area_by_id.get(rack.area_id) if rack.area_id else None
    if not area and row and row.area_id:
        area = layout.area_by_id.get(row.area_id)
    parts = ["Elevations"]
    parts.append(_safe(area.name) if area else "_unassigned")
    parts.append(_safe(row.name if row else rack.row_label or "_no-row"))
    return "/".join(parts) + f"/{_safe(rack.name)}.svg"


def build_office_zip(db: Session, project: Project) -> bytes:
    layout = collect_layout(db, project)
    stem = _safe(project.site_name or project.name, "DCEngineer")
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("How to open in Visio.txt", _readme(layout, stem))
        zf.writestr(f"{stem}.vsdx", build_vsdx(layout))
        zf.writestr(f"{stem} - Data Visualizer.xlsx", build_visualizer_workbook(layout))
        for pic in layout.pictures:
            zf.writestr(pic.zip_path, pic.data)
        for rack in layout.racks:
            devices = layout.devices_by_rack.get(rack.id, [])
            zf.writestr(elevation_zip_path(layout, rack), rack_svg(rack, devices))
    return buf.getvalue()
