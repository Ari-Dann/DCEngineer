"""RBI workbook export (Excel) plus a simple SVG rack elevation."""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models import Cable, Device, Handoff, PDU, PDUPort, Project, Rack


HEADER_FILL = PatternFill("solid", fgColor="1B3A4B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EOL_FILL = PatternFill("solid", fgColor="E85D4C")
NEAR_FILL = PatternFill("solid", fgColor="E8A317")
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def eol_status(eol_date: Optional[str], near_days: int | None = None) -> str:
    if not eol_date:
        return "unknown"
    try:
        parsed = date.fromisoformat(eol_date[:10])
    except ValueError:
        return "unknown"
    today = datetime.now(timezone.utc).date()
    near = near_days if near_days is not None else get_settings().near_eol_days
    if parsed <= today:
        return "eol"
    if parsed <= today + timedelta(days=near):
        return "near"
    return "ok"


def _header(ws, titles: list[str]) -> None:
    for col, title in enumerate(titles, 1):
        cell = ws.cell(1, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(titles))}1"
    ws.freeze_panes = "A2"


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(48, max((len(str(c.value or "")) for c in col), default=10) + 2)
        ws.column_dimensions[letter].width = width


def build_rbi_workbook(db: Session, project: Project) -> bytes:
    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Reliable Baseline Inventory (RBI)"
    cover["A1"].font = Font(size=18, bold=True, color="1B3A4B")
    meta = [
        ("Customer", project.customer),
        ("Project", project.name),
        ("Site", project.site_name),
        ("Address", project.site_address),
        ("Revision", project.revision),
        ("Status", project.status),
        ("Sponsor", project.sponsor),
        ("Start", project.start_date or ""),
        ("Target end", project.target_end_date or ""),
        ("Exported", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (k, v) in enumerate(meta, 3):
        cover.cell(i, 1, k).font = Font(bold=True)
        cover.cell(i, 2, v)
    cover["A14"] = "Photography / data handling"
    cover["A14"].font = Font(bold=True)
    cover["A15"] = project.photography_rules
    cover["A16"] = project.data_handling_rules
    cover["A18"] = "Restricted (government / EMSS) equipment"
    cover["A18"].font = Font(bold=True)
    cover["A19"] = project.restricted_equipment_notes
    cover["A21"] = "Discovery feasibility"
    cover["A21"].font = Font(bold=True)
    cover["A22"] = f"Port access: {project.discovery_port_access}"
    cover["A23"] = f"CDP/LLDP: {project.discovery_cdp_lldp}"
    cover["A24"] = f"SaaS trial: {project.discovery_saas_trial}"
    cover["A25"] = project.discovery_notes
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 60

    rev = wb.create_sheet("Revision Control")
    _header(rev, ["Revision", "Date", "Author", "Notes"])
    rev.append([project.revision, datetime.now(timezone.utc).strftime("%Y-%m-%d"), "", "Exported from DCEngineer"])
    _autosize(rev)

    racks = db.query(Rack).filter(Rack.project_id == project.id).order_by(Rack.name).all()
    rack_sheet = wb.create_sheet("Racks")
    _header(rack_sheet, ["Rack", "Row", "Position", "RU height", "Width in", "Notes"])
    for rack in racks:
        rack_sheet.append([rack.name, rack.row_label, rack.position, rack.ru_height, rack.width_inches, rack.notes])
    _autosize(rack_sheet)

    elev = wb.create_sheet("Elevations")
    _header(
        elev,
        ["Rack", "RU start", "RU end", "Device", "Vendor", "Model", "Type", "Serial", "Function"],
    )
    devices = db.query(Device).filter(Device.project_id == project.id).order_by(Device.rack_id, Device.ru_start).all()
    rack_by_id = {r.id: r for r in racks}
    for dev in devices:
        rack_name = rack_by_id[dev.rack_id].name if dev.rack_id and dev.rack_id in rack_by_id else ""
        elev.append(
            [
                rack_name,
                dev.ru_start,
                dev.ru_end,
                dev.name,
                dev.vendor,
                dev.model,
                dev.device_type,
                dev.serial,
                dev.function,
            ]
        )
    _autosize(elev)

    dev_sheet = wb.create_sheet("Devices")
    _header(
        dev_sheet,
        [
            "Name",
            "Hostname",
            "Rack",
            "Vendor",
            "Model",
            "Serial",
            "Asset",
            "Type",
            "Function",
            "RU start",
            "RU end",
            "Restricted",
            "Restriction",
            "Fan orientation",
            "Mgmt IP",
            "Discovered via",
            "Undocumented",
            "EOL",
            "EOS",
            "EOL status",
            "Notes",
        ],
    )
    for dev in devices:
        rack_name = rack_by_id[dev.rack_id].name if dev.rack_id and dev.rack_id in rack_by_id else ""
        status = eol_status(dev.eol_date)
        row = [
            dev.name,
            dev.hostname,
            rack_name,
            dev.vendor,
            dev.model,
            dev.serial,
            dev.asset_tag,
            dev.device_type,
            dev.function,
            dev.ru_start,
            dev.ru_end,
            "yes" if dev.restricted else "no",
            dev.restricted_reason,
            dev.fan_orientation,
            dev.management_ip,
            dev.discovered_via,
            "yes" if dev.undocumented else "no",
            dev.eol_date or "",
            dev.eos_date or "",
            status,
            dev.notes,
        ]
        dev_sheet.append(row)
        fill = EOL_FILL if status == "eol" else NEAR_FILL if status == "near" else None
        if fill:
            for col in range(1, len(row) + 1):
                dev_sheet.cell(dev_sheet.max_row, col).fill = fill
    _autosize(dev_sheet)

    pdu_sheet = wb.create_sheet("PDU Connectivity")
    _header(pdu_sheet, ["Rack", "PDU", "Bank", "Port", "Device", "Feed", "Amps", "Volts", "Notes"])
    pdus = db.query(PDU).join(Rack).filter(Rack.project_id == project.id).all()
    device_by_id = {d.id: d for d in devices}
    for pdu in pdus:
        rack_name = rack_by_id.get(pdu.rack_id).name if pdu.rack_id in rack_by_id else ""
        ports = db.query(PDUPort).filter(PDUPort.pdu_id == pdu.id).order_by(PDUPort.port_label).all()
        if not ports:
            pdu_sheet.append([rack_name, pdu.name, pdu.bank, "", "", pdu.feed, pdu.amperage, pdu.voltage, ""])
        for port in ports:
            dname = device_by_id[port.device_id].name if port.device_id in device_by_id else ""
            pdu_sheet.append(
                [rack_name, pdu.name, pdu.bank, port.port_label, dname, pdu.feed, pdu.amperage, pdu.voltage, port.notes]
            )
    _autosize(pdu_sheet)

    cab_sheet = wb.create_sheet("Cabling Breakout")
    _header(cab_sheet, ["From", "From port", "To", "To port", "Media", "Color", "Traced", "Notes"])
    cables = db.query(Cable).filter(Cable.project_id == project.id).all()
    for cab in cables:
        cab_sheet.append(
            [cab.from_label, cab.from_port, cab.to_label, cab.to_port, cab.media, cab.color, "yes" if cab.traced else "no", cab.notes]
        )
    _autosize(cab_sheet)

    life = wb.create_sheet("Lifecycle")
    _header(life, ["Device", "Vendor", "Model", "EOL", "EOS", "Status", "Notes"])
    for dev in devices:
        status = eol_status(dev.eol_date)
        life.append([dev.name, dev.vendor, dev.model, dev.eol_date or "", dev.eos_date or "", status, dev.eol_notes])
    _autosize(life)

    rem = wb.create_sheet("Remediation")
    _header(rem, ["Priority", "Device", "Issue", "Recommended action"])
    for dev in devices:
        status = eol_status(dev.eol_date)
        if status == "eol":
            rem.append(["High", dev.name, "Already EOL", "Budget replacement in current cycle"])
        elif status == "near":
            rem.append(["Medium", dev.name, "Near EOL", "Include in next refresh window"])
        if "incorrect" in (dev.fan_orientation or ""):
            rem.append(["Medium", dev.name, f"Fan orientation: {dev.fan_orientation}", "Re-orient or document exception"])
        if dev.undocumented:
            rem.append(["High", dev.name, "Undocumented vs discovery", "Validate ownership and add to CMDB"])
        if dev.restricted:
            rem.append(["Info", dev.name, f"Restricted: {dev.restricted_reason}", "Client-engineer complete remaining fields"])
    _autosize(rem)

    hand = wb.create_sheet("Handoffs")
    _header(hand, ["Date", "From", "To", "Devices captured", "Summary", "Issues", "Follow-ups"])
    for h in db.query(Handoff).filter(Handoff.project_id == project.id).order_by(Handoff.handoff_date).all():
        hand.append([h.handoff_date, h.from_name, h.to_name, h.devices_captured, h.summary, h.issues, h.follow_ups])
    _autosize(hand)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rack_svg(rack: Rack, devices: list[Device]) -> str:
    ru = rack.ru_height or 42
    row_h = 18
    width = 360
    height = ru * row_h + 40
    colors = {
        "server": "#3d9cf0",
        "switch": "#3dbf8c",
        "firewall": "#e87a4c",
        "router": "#5ec8e8",
        "storage": "#9b7dff",
        "pdu": "#e8a317",
        "ups": "#e85d4c",
        "other": "#8b9bb0",
    }
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" width="{width}" height="{height}">',
        f'<rect width="{width}" height="{height}" fill="#0b0f14"/>',
        f'<text x="12" y="22" fill="#e8eef6" font-family="sans-serif" font-size="14">{_esc(rack.name)} — {ru}U</text>',
    ]
    occupied: dict[int, Device] = {}
    for dev in devices:
        if dev.ru_start is None:
            continue
        start = int(dev.ru_start)
        end = int(dev.ru_end or dev.ru_start)
        for u in range(min(start, end), max(start, end) + 1):
            occupied[u] = dev
    for u in range(ru, 0, -1):
        y = 32 + (ru - u) * row_h
        parts.append(
            f'<rect x="40" y="{y}" width="{width-52}" height="{row_h-1}" fill="#121820" stroke="#243040"/>'
        )
        parts.append(
            f'<text x="8" y="{y + 13}" fill="#8b9bb0" font-family="sans-serif" font-size="10">{u}</text>'
        )
        dev = occupied.get(u)
        if dev and (dev.ru_end or dev.ru_start) == u:
            start = int(dev.ru_start)
            end = int(dev.ru_end or dev.ru_start)
            span = abs(end - start) + 1
            top_u = max(start, end)
            y0 = 32 + (ru - top_u) * row_h
            fill = colors.get(dev.device_type, colors["other"])
            parts.append(
                f'<rect x="44" y="{y0 + 1}" width="{width-60}" height="{span * row_h - 3}" rx="3" fill="{fill}" opacity="0.85"/>'
            )
            label = _esc(f"{dev.name}  {dev.vendor} {dev.model}".strip())
            parts.append(
                f'<text x="52" y="{y0 + 12}" fill="#0b0f14" font-family="sans-serif" font-size="11">{label}</text>'
            )
    parts.append("</svg>")
    return "\n".join(parts)


def _esc(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
