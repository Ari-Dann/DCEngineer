"""Import devices and layout (areas / rows / racks) from CSV, XLSX, or ODS."""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from datetime import date, datetime
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.catalog import DEVICE_TYPES, IMPORT_FIELDS, learn_values
from app.layout import apply_row_to_rack
from app.models import AisleRow, Area, Device, Rack

HEADER_MAP = {
    "name": ("name", "device", "device name", "device_name", "unit", "label", "hostname/device"),
    "hostname": ("hostname", "host", "dns", "fqdn"),
    "vendor": ("vendor", "manufacturer", "oem", "make", "mfg"),
    "model": ("model", "part", "pid", "sku", "part number", "part_number", "part no"),
    "serial": ("serial", "serial number", "serial_number", "sn", "s/n", "s/n."),
    "asset_tag": ("asset", "asset tag", "asset_tag", "tag", "asset no"),
    "rack": ("rack", "rack name", "rack_name", "cabinet", "cab"),
    "row": ("row", "aisle", "row name", "row_name", "aisle name"),
    "area": ("area", "hall", "cage", "room", "area name"),
    "ru_start": ("ru start", "ru_start", "ru", "u", "u start", "position", "ru position"),
    "ru_end": ("ru end", "ru_end", "u end"),
    "ru_height": ("height", "height u", "ru height", "ru_height", "u height"),
    "device_type": ("type", "device type", "device_type", "class", "category"),
    "function": ("function", "role", "purpose"),
    "management_ip": ("ip", "mgmt ip", "management_ip", "mgmt", "management ip"),
    "notes": ("notes", "note", "comment", "comments"),
    "eol_date": ("eol", "eol date", "eol_date", "end of life"),
    "eos_date": ("eos", "eos date", "eos_date", "end of sale", "end of support"),
    "fan_orientation": ("fan", "fan orientation", "fan_orientation", "airflow"),
    "indicator_type": ("led", "screen", "display", "indicator", "led screen", "led/screen"),
    "indicator_color": ("led color", "screen color", "indicator color", "color", "light color"),
}

KNOWN_FIELDS = {f["id"] for f in IMPORT_FIELDS}
PREFERRED_SHEETS = ("devices", "device list", "inventory", "assets", "equipment", "elevations")
GENERIC_SHEET_NAMES = {
    *PREFERRED_SHEETS,
    "cover",
    "revision control",
    "racks",
    "pdu connectivity",
    "cabling",
    "lifecycle",
    "remediation",
    "handoffs",
    "layout",
    "sheet1",
    "sheet2",
    "sheet3",
}
KNOWN_TYPES = {t.lower() for t in DEVICE_TYPES}
_BLANK_TEXT = {"", "unknown", "n/a", "na", "none", "-"}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _label_key(cell: Any) -> str:
    text = _norm(cell).lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def known_field(cell: Any) -> str | None:
    text = _label_key(cell)
    if not text:
        return None
    for field, aliases in HEADER_MAP.items():
        if text == field.replace("_", " ") or text in aliases:
            return field
    return None


def _mapped_count(cells: list[str]) -> int:
    seen: set[str] = set()
    for cell in cells:
        field = known_field(cell)
        if field:
            seen.add(field)
    return len(seen)


def _grid_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [[_norm(c) for c in row] for row in reader]


def _sheets_from_xlsx(data: bytes) -> list[tuple[str, list[list[str]]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        grid = []
        for raw in ws.iter_rows(values_only=True):
            row = [_norm(c) for c in raw]
            if any(row):
                grid.append(row)
        out.append((ws.title, grid))
    return out


_ODS_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def _is_ods(data: bytes) -> bool:
    if data[:2] != b"PK":
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = set(zf.namelist())
            if "mimetype" in names:
                mime = zf.read("mimetype").decode("utf-8", errors="ignore")
                if "opendocument.spreadsheet" in mime:
                    return True
            return "content.xml" in names and "[Content_Types].xml" not in names
    except zipfile.BadZipFile:
        return False


def _ods_cell_text(cell: ET.Element) -> str:
    value = cell.get(f"{{{_ODS_NS['office']}}}value")
    if value:
        return _norm(value)
    parts = ["".join(p.itertext()) for p in cell.findall("text:p", _ODS_NS)]
    return _norm("\n".join(parts) if parts else "")


def _sheets_from_ods(data: bytes) -> list[tuple[str, list[list[str]]]]:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        root = ET.fromstring(zf.read("content.xml"))
    out: list[tuple[str, list[list[str]]]] = []
    for table in root.findall(".//table:table", _ODS_NS):
        name = table.get(f"{{{_ODS_NS['table']}}}name") or "Sheet1"
        grid: list[list[str]] = []
        for row_el in table.findall("table:table-row", _ODS_NS):
            cells: list[str] = []
            for cell in row_el.findall("table:table-cell", _ODS_NS):
                text = _ods_cell_text(cell)
                repeat = min(int(cell.get(f"{{{_ODS_NS['table']}}}number-columns-repeated") or 1), 256)
                cells.extend([text] * repeat)
            while cells and not cells[-1]:
                cells.pop()
            row_repeat = int(row_el.get(f"{{{_ODS_NS['table']}}}number-rows-repeated") or 1)
            if not any(cells):
                continue
            for _ in range(min(row_repeat, 50)):
                grid.append(cells)
        if grid:
            out.append((name, grid))
    return out


def parse_workbook(filename: str, data: bytes) -> list[tuple[str, list[list[str]]]]:
    name = (filename or "").lower()
    if name.endswith(".xls"):
        raise ValueError("Legacy .xls is not supported. Save as .xlsx, .ods, or .csv and try again.")
    if name.endswith(".csv") or name.endswith(".txt"):
        return [(filename or "Sheet1", _grid_from_csv(data))]
    if name.endswith(".ods") or _is_ods(data):
        sheets = _sheets_from_ods(data)
        if not sheets:
            raise ValueError("ODS workbook contained no tables.")
        return sheets
    if name.endswith(".xlsx") or data[:2] == b"PK":
        return _sheets_from_xlsx(data)
    return [(filename or "Sheet1", _grid_from_csv(data))]


def _best_header_index(grid: list[list[str]], orientation: str) -> int:
    best_i, best_score = 0, -1
    if orientation == "rows":
        for i, row in enumerate(grid[:12]):
            score = _mapped_count(row)
            if score > best_score:
                best_i, best_score = i, score
        return best_i
    width = max((len(r) for r in grid), default=0)
    for i in range(min(5, width)):
        labels = [r[i] if i < len(r) else "" for r in grid]
        score = _mapped_count(labels)
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def detect_orientation(grid: list[list[str]]) -> str:
    if not grid:
        return "rows"
    row_score = max((_mapped_count(row) for row in grid[:12]), default=0)
    col_score = _mapped_count([r[0] if r else "" for r in grid])
    return "columns" if col_score > row_score else "rows"


def suggest_mapping(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        field = known_field(header)
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def _trim_trailing(row: list[str]) -> list[str]:
    out = list(row)
    while out and not out[-1]:
        out.pop()
    return out


def _headers_for(grid: list[list[str]], orientation: str, header_index: int) -> list[str]:
    if orientation == "rows":
        return _trim_trailing(list(grid[header_index]) if header_index < len(grid) else [])
    return _trim_trailing([r[header_index] if header_index < len(r) else "" for r in grid])


def records_from_grid(
    grid: list[list[str]],
    orientation: str,
    header_index: int,
    mapping: dict[str, int] | None,
) -> list[dict[str, str]]:
    headers = _headers_for(grid, orientation, header_index)
    used = mapping if mapping is not None else suggest_mapping(headers)
    records: list[dict[str, str]] = []
    if orientation == "rows":
        for row in grid[header_index + 1 :]:
            rec = {field: (row[idx] if idx < len(row) else "") for field, idx in used.items()}
            if any(_norm(v) for v in rec.values()):
                records.append(rec)
        return records
    width = max((len(r) for r in grid), default=0)
    for col in range(header_index + 1, width):
        rec = {}
        for field, row_idx in used.items():
            row = grid[row_idx] if row_idx < len(grid) else []
            rec[field] = row[col] if col < len(row) else ""
        if any(_norm(v) for v in rec.values()):
            records.append(rec)
    return records


def _raw_sample(grid: list[list[str]], orientation: str, header_index: int, limit: int = 8) -> list[list[str]]:
    if orientation == "rows":
        return [list(r) for r in grid[header_index + 1 : header_index + 1 + limit]]
    width = max((len(r) for r in grid), default=0)
    sample = []
    for col in range(header_index + 1, min(width, header_index + 1 + limit)):
        sample.append([r[col] if col < len(r) else "" for r in grid])
    return sample


def describe_sheet(
    name: str,
    grid: list[list[str]],
    *,
    orientation: str | None = None,
    header_index: int | None = None,
    mapping: dict[str, int] | None = None,
) -> dict[str, Any]:
    orientation = orientation or detect_orientation(grid)
    if header_index is None:
        header_index = _best_header_index(grid, orientation)
    headers = _headers_for(grid, orientation, header_index)
    suggested = mapping if mapping is not None else suggest_mapping(headers)
    records = records_from_grid(grid, orientation, header_index, suggested)
    return {
        "name": name,
        "orientation": orientation,
        "header_index": header_index,
        "headers": [
            {"index": i, "label": label or f"Column {i + 1}", "suggested": suggested_field}
            for i, label in enumerate(headers)
            for suggested_field in [next((f for f, idx in suggested.items() if idx == i), "")]
        ],
        "raw_sample": _raw_sample(grid, orientation, header_index),
        "sample_records": records[:8],
        "record_count": len(records),
        "mapped_fields": sorted(suggested.keys()),
    }


def pick_sheet(sheets: list[dict[str, Any]]) -> str:
    if not sheets:
        return ""
    by_name = {s["name"].lower().strip(): s for s in sheets}
    for pref in PREFERRED_SHEETS:
        if pref in by_name:
            return by_name[pref]["name"]
    return max(sheets, key=lambda s: (len(s.get("mapped_fields") or []), s.get("record_count") or 0))["name"]


def preview_import(
    filename: str,
    data: bytes,
    *,
    sheet: str | None = None,
    orientation: str | None = None,
    header_index: int | None = None,
) -> dict[str, Any]:
    parsed = parse_workbook(filename, data)
    if not parsed:
        return {"filename": filename, "sheets": [], "suggested_sheet": "", "fields": IMPORT_FIELDS}
    described = []
    for name, grid in parsed:
        use_opts = name == sheet if sheet else False
        described.append(
            describe_sheet(
                name,
                grid,
                orientation=orientation if use_opts or len(parsed) == 1 else None,
                header_index=header_index if use_opts or len(parsed) == 1 else None,
            )
        )
    suggested = sheet or pick_sheet(described)
    return {
        "filename": filename,
        "sheets": described,
        "suggested_sheet": suggested,
        "fields": IMPORT_FIELDS,
    }


def _int(value: str) -> int | None:
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def _normalize_type(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return "server"
    lower = text.lower()
    if lower in KNOWN_TYPES:
        return lower
    return text


def _parse_mapping(raw: str | dict | None) -> dict[str, int] | None:
    if not raw:
        return None
    data = json.loads(raw) if isinstance(raw, str) else raw
    out: dict[str, int] = {}
    for key, value in (data or {}).items():
        if key not in KNOWN_FIELDS or value is None or value == "":
            continue
        try:
            out[key] = int(value)
        except (TypeError, ValueError):
            continue
    return out or None


def _empty_result(errors: list[str] | None = None) -> dict:
    return {
        "created": 0,
        "updated": 0,
        "racks_created": 0,
        "areas_created": 0,
        "rows_created": 0,
        "preserved": 0,
        "skipped": 0,
        "rows": 0,
        "errors": errors or [],
        "names": [],
        "sheet": "",
        "orientation": "rows",
    }


def _is_generic_sheet(name: str) -> bool:
    label = (name or "").strip().lower()
    if label in GENERIC_SHEET_NAMES:
        return True
    # CSV/TXT uploads use the filename as the sheet title; that is not an area or row.
    return label.endswith((".csv", ".txt", ".xlsx", ".ods", ".xls"))


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip().lower() in _BLANK_TEXT:
        return False
    return True


def _sheet_hierarchy_hint(sheet_name: str) -> str:
    label = (sheet_name or "").strip()
    if not label or _is_generic_sheet(label):
        return ""
    return label


def _path_label(area: Area | None, aisle: AisleRow | None, rack: Rack | None) -> str:
    return " / ".join(part for part in (area.name if area else "", aisle.name if aisle else "", rack.name if rack else "") if part)


class _HierarchyIndex:
    """Match import records to existing layout without stealing items from another parent."""

    def __init__(self, db: Session, project_id: int, default_area: Area | None):
        self.db = db
        self.project_id = project_id
        self.default_area = default_area
        self.areas_created = 0
        self.rows_created = 0
        self.racks_created = 0
        self.areas: dict[str, Area] = {
            a.name.lower(): a for a in db.query(Area).filter(Area.project_id == project_id).all()
        }
        self.rows: list[AisleRow] = db.query(AisleRow).filter(AisleRow.project_id == project_id).all()
        self.racks: list[Rack] = db.query(Rack).filter(Rack.project_id == project_id).all()
        self.devices: list[Device] = db.query(Device).filter(Device.project_id == project_id).all()

    def area(self, name: str) -> Area | None:
        label = (name or "").strip()
        if not label:
            return self.default_area
        key = label.lower()
        found = self.areas.get(key)
        if found:
            return found
        found = Area(project_id=self.project_id, name=label)
        self.db.add(found)
        self.db.flush()
        self.areas[key] = found
        self.areas_created += 1
        return found

    def row(self, name: str, area: Area | None) -> AisleRow | None:
        label = (name or "").strip()
        if not label:
            return None
        key = label.lower()
        if area:
            for row in self.rows:
                if row.area_id == area.id and row.name.lower() == key:
                    return row
        else:
            for row in self.rows:
                if row.area_id is None and row.name.lower() == key:
                    return row
        row = AisleRow(project_id=self.project_id, area_id=area.id if area else None, name=label)
        self.db.add(row)
        self.db.flush()
        self.rows.append(row)
        self.rows_created += 1
        return row

    def rack(self, name: str, aisle: AisleRow | None, area: Area | None) -> Rack | None:
        label = (name or "").strip()
        if not label:
            return None
        key = label.lower()
        area_id = area.id if area else None
        named = [rack for rack in self.racks if rack.name.lower() == key]
        found: Rack | None = None
        if aisle:
            in_row = [rack for rack in named if rack.row_id == aisle.id]
            if in_row:
                found = in_row[0]
            else:
                unassigned = [
                    rack
                    for rack in named
                    if rack.row_id is None
                    and (rack.area_id == area_id or (rack.area_id is None and area_id is None))
                ]
                if unassigned:
                    found = unassigned[0]
                    apply_row_to_rack(found, aisle, area_id)
        elif area_id:
            in_area = [rack for rack in named if rack.area_id == area_id]
            if len(in_area) == 1:
                found = in_area[0]
            else:
                unassigned = [rack for rack in in_area if rack.row_id is None]
                if len(unassigned) == 1:
                    found = unassigned[0]
        else:
            if len(named) == 1:
                found = named[0]
            else:
                unassigned = [rack for rack in named if rack.row_id is None and rack.area_id is None]
                if len(unassigned) == 1:
                    found = unassigned[0]
        if found:
            return found
        rack = Rack(
            project_id=self.project_id,
            name=label,
            ru_height=42,
            area_id=aisle.area_id if aisle and aisle.area_id is not None else area_id,
            row_id=aisle.id if aisle else None,
            row_label=aisle.name if aisle else "",
        )
        self.db.add(rack)
        self.db.flush()
        self.racks.append(rack)
        self.racks_created += 1
        return rack

    def device(self, *, serial: str, name: str, rack: Rack | None) -> Device | None:
        serial_key = (serial or "").strip().lower()
        name_key = (name or "").strip().lower()
        if rack and serial_key:
            for device in self.devices:
                if device.rack_id == rack.id and (device.serial or "").strip().lower() == serial_key:
                    return device
        if rack and name_key:
            for device in self.devices:
                if device.rack_id == rack.id and (device.name or "").strip().lower() == name_key:
                    existing_serial = (device.serial or "").strip().lower()
                    if serial_key and existing_serial and existing_serial != serial_key:
                        continue
                    return device
        if serial_key:
            for device in self.devices:
                if (device.serial or "").strip().lower() == serial_key:
                    return device
        return None

    def remember(self, device: Device) -> None:
        self.devices.append(device)


def _hierarchy_for_record(
    index: _HierarchyIndex,
    row: dict[str, str],
    sheet_name: str,
) -> tuple[Area | None, AisleRow | None, Rack | None]:
    explicit_area = (row.get("area") or "").strip()
    explicit_row = (row.get("row") or "").strip()
    rack_name = (row.get("rack") or "").strip()
    hint = _sheet_hierarchy_hint(sheet_name)
    if explicit_area:
        area = index.area(explicit_area)
        row_name = explicit_row or (hint if hint.lower() != explicit_area.lower() else "")
    elif index.default_area:
        area = index.default_area
        row_name = explicit_row or hint
    elif hint:
        area = index.area(hint)
        row_name = explicit_row
    else:
        area = None
        row_name = explicit_row
    aisle = index.row(row_name, area) if row_name else None
    rack = index.rack(rack_name, aisle, area) if rack_name else None
    return area, aisle, rack


def _device_payload(row: dict[str, str], name: str, rack: Rack | None) -> dict[str, Any]:
    ru_start = _int(row.get("ru_start", ""))
    ru_end = _int(row.get("ru_end", ""))
    height = _int(row.get("ru_height", ""))
    if ru_start is not None and ru_end is None and height:
        ru_end = ru_start + height - 1
    serial = (row.get("serial") or "").strip()
    payload: dict[str, Any] = {
        "name": name[:255],
        "serial": serial[:128],
        "discovered_via": "import",
    }
    if rack:
        payload["rack_id"] = rack.id
    mapped = {
        "hostname": (row.get("hostname") or "")[:255],
        "vendor": (row.get("vendor") or "")[:128],
        "model": (row.get("model") or "")[:128],
        "asset_tag": (row.get("asset_tag") or "")[:128],
        "function": (row.get("function") or "")[:255],
        "management_ip": (row.get("management_ip") or "")[:64],
        "notes": row.get("notes") or "",
        "eol_date": row.get("eol_date") or None,
        "eos_date": row.get("eos_date") or None,
        "fan_orientation": (row.get("fan_orientation") or "")[:64],
        "indicator_type": (row.get("indicator_type") or "")[:32],
        "indicator_color": (row.get("indicator_color") or "")[:32],
    }
    for key, value in mapped.items():
        if _has_value(value):
            payload[key] = value
    if _has_value(row.get("device_type")):
        payload["device_type"] = _normalize_type(row.get("device_type", ""))[:64]
    if ru_start is not None:
        payload["ru_start"] = ru_start
    if ru_end is not None:
        payload["ru_end"] = ru_end
    return payload


def _apply_payload(device: Device, payload: dict[str, Any], *, allow_rack: bool) -> None:
    for key, value in payload.items():
        if key == "discovered_via":
            continue
        if key == "rack_id" and not allow_rack:
            continue
        if not _has_value(value) and value not in (0, False):
            continue
        setattr(device, key, value)


def import_devices(
    db: Session,
    project_id: int,
    filename: str,
    data: bytes,
    user_id: int | None,
    *,
    sheet: str | None = None,
    orientation: str | None = None,
    header_index: int | None = None,
    mapping: dict[str, int] | str | None = None,
    default_area_id: int | None = None,
    all_sheets: bool = False,
) -> dict:
    parsed = parse_workbook(filename, data)
    if not parsed:
        return _empty_result(["File contained no rows"])

    described = [describe_sheet(n, g) for n, g in parsed]
    chosen = sheet or pick_sheet(described)
    default_area = None
    if default_area_id:
        default_area = db.get(Area, default_area_id)
        if not default_area or default_area.project_id != project_id:
            return _empty_result(["Default area was not found in this project"])

    if all_sheets:
        targets = [(n, g) for n, g in parsed if any(any(cell for cell in row) for row in g)]
    else:
        grid = next((g for n, g in parsed if n == chosen), parsed[0][1])
        used_sheet = next((n for n, g in parsed if n == chosen), parsed[0][0])
        targets = [(used_sheet, grid)]

    used_mapping = _parse_mapping(mapping)
    index = _HierarchyIndex(db, project_id, default_area)
    created = 0
    updated = 0
    preserved = 0
    skipped = 0
    errors: list[str] = []
    names: list[str] = []
    total_rows = 0
    orientations: list[str] = []
    sheet_names: list[str] = []

    for sheet_name, grid in targets:
        sheet_orientation = orientation if (not all_sheets or sheet_name == chosen) else None
        sheet_orientation = sheet_orientation or detect_orientation(grid)
        sheet_header = header_index if (not all_sheets or sheet_name == chosen) else None
        if sheet_header is None:
            sheet_header = _best_header_index(grid, sheet_orientation)
        sheet_mapping = used_mapping if (not all_sheets or sheet_name == chosen) else None
        records = records_from_grid(grid, sheet_orientation, sheet_header, sheet_mapping)
        if not records:
            continue
        total_rows += len(records)
        orientations.append(sheet_orientation)
        sheet_names.append(sheet_name)

        # Pass 1: Area → Row → Rack so later device rows never re-parent populated layout.
        for row in records:
            if not any((row.get("area"), row.get("row"), row.get("rack"), row.get("name"), row.get("hostname"), row.get("serial"))):
                continue
            _hierarchy_for_record(index, row, sheet_name)

        # Pass 2: devices, matched only after hierarchy is resolved.
        for offset, row in enumerate(records, start=2):
            name = (row.get("name") or row.get("hostname") or row.get("serial") or "").strip()
            area_name = (row.get("area") or "").strip()
            row_name = (row.get("row") or "").strip()
            rack_name = (row.get("rack") or "").strip()
            if not name and not area_name and not row_name and not rack_name:
                skipped += 1
                continue
            _area, _aisle, rack = _hierarchy_for_record(index, row, sheet_name)
            if not name:
                continue
            ru_end = _int(row.get("ru_end", ""))
            height = _int(row.get("ru_height", ""))
            ru_start = _int(row.get("ru_start", ""))
            if ru_start is not None and ru_end is None and height:
                ru_end = ru_start + height - 1
            if rack and ru_end and ru_end > rack.ru_height:
                rack.ru_height = min(70, ru_end)

            payload = _device_payload(row, name, rack)
            existing = index.device(serial=payload.get("serial") or "", name=name, rack=rack)
            try:
                if existing:
                    located_elsewhere = bool(rack and existing.rack_id and existing.rack_id != rack.id)
                    if located_elsewhere:
                        current_rack = next((r for r in index.racks if r.id == existing.rack_id), None)
                        current_area = next(
                            (a for a in index.areas.values() if current_rack and a.id == current_rack.area_id),
                            None,
                        )
                        current_row = next(
                            (r for r in index.rows if current_rack and r.id == current_rack.row_id),
                            None,
                        )
                        preserved += 1
                        errors.append(
                            f"{sheet_name} row {offset}: serial already at {_path_label(current_area, current_row, current_rack)}; left in place"
                        )
                        continue
                    allow_rack = not rack or existing.rack_id in (None, rack.id)
                    _apply_payload(existing, payload, allow_rack=allow_rack)
                    if allow_rack and rack and existing.rack_id is None:
                        existing.rack_id = rack.id
                    updated += 1
                else:
                    defaults = dict(
                        hostname="",
                        vendor="",
                        model="",
                        asset_tag="",
                        device_type="server",
                        function="",
                        management_ip="",
                        notes="",
                        fan_orientation="unknown",
                        indicator_type="unknown",
                        indicator_color="unknown",
                    )
                    defaults.update(payload)
                    device = Device(project_id=project_id, captured_by=user_id, **defaults)
                    db.add(device)
                    db.flush()
                    index.remember(device)
                    created += 1
                names.append(payload["name"])
                learn_values(
                    db,
                    vendor=payload.get("vendor") or "",
                    model=payload.get("model") or "",
                    device_type=payload.get("device_type") or "",
                    function=payload.get("function") or "",
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{sheet_name} row {offset}: {exc}")
                skipped += 1

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "racks_created": index.racks_created,
        "areas_created": index.areas_created,
        "rows_created": index.rows_created,
        "preserved": preserved,
        "skipped": skipped,
        "rows": total_rows,
        "errors": errors[:20],
        "names": names[:25],
        "sheet": ", ".join(sheet_names),
        "orientation": orientations[0] if orientations else "rows",
    }


# Back-compat for older tests / callers that only need row dicts from a simple table.
def parse_table(filename: str, data: bytes) -> list[dict[str, str]]:
    preview = preview_import(filename, data)
    if not preview["sheets"]:
        return []
    chosen = next((s for s in preview["sheets"] if s["name"] == preview["suggested_sheet"]), preview["sheets"][0])
    parsed = parse_workbook(filename, data)
    grid = next((g for n, g in parsed if n == chosen["name"]), parsed[0][1])
    mapping = {h["suggested"]: h["index"] for h in chosen["headers"] if h["suggested"]}
    return records_from_grid(grid, chosen["orientation"], chosen["header_index"], mapping or None)
